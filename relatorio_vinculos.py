# -*- coding: utf-8 -*-
"""Relatório de PROCESSOS VINCULADOS (têm terceiro interessado e/ou fundo).
Mesmo padrão consolidado (1 linha por CPF, ADM+JUD+FIN), focado nos vínculos.
Inclui um CPF se QUALQUER card dele (ADM/JUD/FIN) tiver terceiro/fundo preenchido.
Saída: ../PROCESSOS_VINCULADOS.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
SAIDA=os.path.join(os.path.dirname(__file__),"..","PROCESSOS_VINCULADOS.xlsx")

FID={"ADM":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio_1","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "JUD":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interassado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "FIN":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=FID[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
adm=pull(ADM,"ADM"); jud=pull(JUD,"JUD"); fin=pull(FIN,"FIN")
print(f"  ADM {len(adm)} | JUD {len(jud)} | FIN {len(fin)}")
by=defaultdict(lambda:{"ADM":[],"JUD":[],"FIN":[]})
for c in adm+jud+fin:
    if c["cpf"]: by[c["cpf"]][c["kind"]].append(c)

def pick(cards, field):
    for c in cards:
        if c[field]: return c[field]
    return ""
def nome_of(d):
    for k in ("ADM","JUD","FIN"):
        for c in d[k]:
            if c["nome"]: return c["nome"]
    return ""

registros=[]
for cpf,d in by.items():
    terc_adm=pick(d["ADM"],"terc"); terc_jud=pick(d["JUD"],"terc"); terc_fin=pick(d["FIN"],"terc")
    inv=pick(d["ADM"]+d["JUD"]+d["FIN"],"inv")
    esp=pick(d["ADM"]+d["JUD"]+d["FIN"],"esp")
    if not (terc_adm or terc_jud or terc_fin or inv or esp):
        continue  # sem vínculo -> fora
    # terceiro consolidado (valor predominante)
    tvals=[t for t in (terc_adm,terc_jud,terc_fin) if t]
    terc=Counter(tvals).most_common(1)[0][0] if tvals else ""
    pipes="+".join(k for k in ("ADM","JUD","FIN") if d[k])
    registros.append({
        "Nome": nome_of(d), "CPF": fcpf(cpf),
        "Terceiro Interessado": terc,
        "Terceiro (ADM)": terc_adm, "Terceiro (JUD)": terc_jud, "Terceiro (FIN)": terc_fin,
        "Fundo Investidor": inv, "Fundo Especial": esp,
        "Pipes": pipes,
    })

registros.sort(key=lambda x:(x["Terceiro Interessado"] or "zzz", x["Fundo Especial"], str(x["Nome"])))
print(f"\nProcessos vinculados: {len(registros)}")

# ── Excel ──
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Vínculos"
headers=["Nome","CPF","Terceiro Interessado","Terceiro (ADM)","Terceiro (JUD)","Terceiro (FIN)","Fundo Investidor","Fundo Especial","Pipes"]
ws.append(headers)
for r in registros: ws.append([r[h] for h in headers])
font=Font(bold=True,color="FFFFFF"); thin=Side(style="thin",color="BFBFBF")
col_fill={1:"1F4E78",2:"1F4E78",3:"7C3AED",4:"7C3AED",5:"7C3AED",6:"7C3AED",7:"548235",8:"548235",9:"404040"}
for i,h in enumerate(headers,1):
    c=ws.cell(1,i); c.fill=PatternFill("solid",fgColor=col_fill.get(i,"404040")); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=thin)
ws.freeze_panes="A2"; ws.row_dimensions[1].height=30
for i,w in enumerate([34,15,28,24,24,24,22,18,12],1):
    ws.column_dimensions[get_column_letter(i)].width=w

# ── Resumo ──
ws2=wb.create_sheet("Resumo")
def bloco(titulo, counter, start_row):
    ws2.cell(start_row,1,titulo).font=Font(bold=True,color="FFFFFF")
    ws2.cell(start_row,1).fill=PatternFill("solid",fgColor="1F4E78")
    ws2.cell(start_row,2).fill=PatternFill("solid",fgColor="1F4E78")
    r=start_row+1
    for k,v in counter.most_common():
        ws2.cell(r,1,k or "(vazio)"); ws2.cell(r,2,v); r+=1
    return r+1

# terceiro consolidado por CPF
terc_counter=Counter(r["Terceiro Interessado"] for r in registros if r["Terceiro Interessado"])
esp_counter =Counter(r["Fundo Especial"] for r in registros if r["Fundo Especial"])
inv_counter =Counter(r["Fundo Investidor"] for r in registros if r["Fundo Investidor"])
ws2.append(["Métrica","Valor"]);
for c in ws2[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="404040")
ws2.append(["Total processos vinculados", len(registros)])
ws2.append(["Com Terceiro Interessado", sum(1 for r in registros if r['Terceiro Interessado'])])
ws2.append(["Com Fundo Especial", sum(1 for r in registros if r['Fundo Especial'])])
ws2.append(["Com Fundo Investidor", sum(1 for r in registros if r['Fundo Investidor'])])
nr=bloco("POR TERCEIRO INTERESSADO", terc_counter, 7)
nr=bloco("POR FUNDO ESPECIAL", esp_counter, nr)
nr=bloco("POR FUNDO INVESTIDOR", inv_counter, nr)
ws2.column_dimensions["A"].width=46; ws2.column_dimensions["B"].width=12

wb.save(SAIDA)
print("Salvo:",os.path.abspath(SAIDA))
print("\nTop terceiros:")
for k,v in terc_counter.most_common(12): print(f"  {v:4}  {k}")
