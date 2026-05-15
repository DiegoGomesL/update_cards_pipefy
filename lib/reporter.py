from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOGS_DIR = Path(__file__).parent.parent / "logs"

COLUMNS = [
    ("card_id",   "ID do Card",       14),
    ("nome",      "Nome",             40),
    ("old_value", "Valor Anterior",   25),
    ("new_value", "Valor Novo",       25),
    ("status",    "Status",           14),
    ("error",     "Erro",             35),
    ("timestamp", "Data/Hora",        20),
]

# Cores por status
STATUS_FILLS = {
    "OK":           PatternFill("solid", fgColor="C6EFCE"),  # verde claro
    "FALHA":        PatternFill("solid", fgColor="FFC7CE"),  # vermelho claro
    "DRY-RUN":      PatternFill("solid", fgColor="FFEB9C"),  # amarelo claro
    "INTERROMPIDO": PatternFill("solid", fgColor="FFEB9C"),  # amarelo claro
}
STATUS_FONTS = {
    "OK":           Font(bold=True, color="276221"),
    "FALHA":        Font(bold=True, color="9C0006"),
    "DRY-RUN":      Font(bold=True, color="7D6608"),
    "INTERROMPIDO": Font(bold=True, color="7D6608"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


def save(
    results: list[dict],
    pipe_label: str,
    field_id: str,
    dry_run: bool = False,
) -> Path:
    """
    Salva resultados em .xlsx formatado no diretório logs/.
    Retorna o Path do arquivo gerado.
    """
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    now       = datetime.now()
    suffix    = "_DRYRUN" if dry_run else ""
    filename  = f"{now.strftime('%Y-%m-%d_%H-%M')}_{pipe_label}_{field_id}{suffix}.xlsx"
    filepath  = LOGS_DIR / filename
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    # ── Cabeçalho ────────────────────────────────────────────
    for col_idx, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"  # congela cabeçalho

    # ── Dados ─────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=2):
        status = r.get("status", "")
        row_fill = STATUS_FILLS.get(status)

        values = [
            r.get("card_id",   ""),
            r.get("nome",      ""),
            r.get("old_value", ""),
            r.get("new_value", ""),
            status,
            r.get("error",     ""),
            timestamp,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Coluna status: cor e negrito específicos
            if col_idx == 5 and status in STATUS_FILLS:
                cell.fill = STATUS_FILLS[status]
                cell.font = STATUS_FONTS[status]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_fill and row_idx % 2 == 0:
                pass  # linhas pares ficam brancas para facilitar leitura

            # Linhas alternadas em cinza bem claro
            if col_idx != 5 and row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # ── Filtro automático ─────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    return filepath


def summary(results: list[dict]) -> dict:
    """Retorna contagens por status."""
    counts: dict[str, int] = {}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    return counts
