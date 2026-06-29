# -*- coding: utf-8 -*-
"""Lista dos ~900 processos do fundo WAIMEA, 1 linha por processo (CPF), com:
Nome | CPF | Data Prot. ADM | Nº Prot. ADM | Resultado ADM | Data Prot. JUD | Nº Prot. JUD | Resultado JUD
Universo = carteira Waimea (WAIMEA_900_Processos.xlsx); dados ADM/JUD do Pipefy.
Saída: ../WAIMEA_900_PROTOCOLOS.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]
CARTEIRA=os.path.join(os.path.dirname(__file__),"..","WAIMEA_900_Processos.xlsx")
SAIDA=os.path.join(os.path.dirname(__file__),"..","WAIMEA_900_PROTOCOLOS.xlsx")

F={"ADM":{"cpf":"cpf_do_benefici_rio_1","nome":"nome_do_benefici_rio",
          "prot":"n_mero_do_processo_administrativo","data":"data_do_protocolo","res":"resultado_do_pedido"},
   "JUD":{"cpf":"cpf_do_benefici_rio","nome":"nome_do_benefici_rio",
          "prot":"n_mero_do_processo","data":"data_do_protocolo","res":"resultado_do_processo"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS"}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c
def fdata(v):
    m=re.match(r"(\d{4})-(\d{2})-(\d{2})",nz(v)); return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else nz(v)
def cap(s):
    s=nz(s); return s[:1].upper()+s[1:].lower() if s and s.isupper() else s

# carteira Waimea (universo de ~900)
wb=openpyxl.load_workbook(CARTEIRA,read_only=True,data_only=True); ws=wb["Processos"]
rows=list(ws.iter_rows(values_only=True)); wb.close()
ordem=[]; waimea_nome={}
for r in rows[1:]:
    c=cpf11(r[9])
    if c and len(c)==11 and c not in waimea_nome:
        waimea_nome[c]=nz(r[7]); ordem.append(c)
print(f"CPFs Waimea na carteira: {len(waimea_nome)}")

# Complemento: 'processos waimea.xlsx' (dados ADM de casos que estavam sem informação)
SUPP=os.path.join(os.path.dirname(__file__),"..","processos waimea.xlsx")
supp={}
if os.path.exists(SUPP):
    sw=openpyxl.load_workbook(SUPP,read_only=True,data_only=True); sws=sw.active
    for r in list(sws.iter_rows(min_row=3,values_only=True)):
        c=cpf11(r[1])
        if not (c and len(c)==11): continue
        supp[c]={"nome":nz(r[0]),"numADM":nz(r[4]),"dataADM":r[5],"resADM":nz(r[7]),
                 "numJUD":nz(r[8]),"dataJUD":r[9],"resJUD":nz(r[11])}
    sw.close()
    # adiciona ao universo os CPFs do complemento que não estão na carteira
    novos_supp=[c for c in supp if c not in waimea_nome]
    for c in novos_supp:
        waimea_nome[c]=supp[c]["nome"]; ordem.append(c)
    print(f"Complemento: {len(supp)} casos | novos fora da carteira: {len(novos_supp)} | universo agora: {len(ordem)}")

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=F[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "prot":nz(fm.get(fi["prot"])),"data":nz(fm.get(fi["data"])),"res":nz(fm.get(fi["res"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy (ADM/JUD)...")
adm_by=defaultdict(list); jud_by=defaultdict(list)
for c in pull(ADM,"ADM"):
    if c["cpf"] in waimea_nome: adm_by[c["cpf"]].append(c)
for c in pull(JUD,"JUD"):
    if c["cpf"] in waimea_nome: jud_by[c["cpf"]].append(c)

def is_active(c,term): return c["phase"].upper() not in term
def pick(cards, term):
    """prefere card COM resultado; senão ativo; senão o mais recente."""
    if not cards: return None
    com_res=[c for c in cards if c["res"]]
    if com_res: return max(com_res,key=lambda c:(c["data"],c["prot"]))
    ativos=[c for c in cards if is_active(c,term)]
    pool=ativos if ativos else cards
    return max(pool,key=lambda c:(c["data"],c["prot"]))

registros=[]
n_pipefy=0; n_complemento=0
for cpf in ordem:
    a=pick(adm_by.get(cpf,[]),ADM_TERM); j=pick(jud_by.get(cpf,[]),JUD_TERM)
    tem_card = bool(adm_by.get(cpf) or jud_by.get(cpf))
    if a or j: n_pipefy+=1
    nome=(a or j or {}).get("nome") or waimea_nome.get(cpf,"")
    data_adm=fdata(a["data"]) if a else ""; num_adm=a["prot"] if a else ""
    res_adm=cap(a["res"]) if a else ""
    data_jud=fdata(j["data"]) if j else ""; num_jud=j["prot"] if j else ""
    res_jud=cap(j["res"]) if j else ""
    origem="Pipefy" if (a or j) else ""

    # Complemento: preenche ADM (e JUD se houver) quando o Pipefy não tem
    s=supp.get(cpf)
    if s:
        if not num_adm and s["numADM"]:
            num_adm=s["numADM"]; data_adm=fdata(s["dataADM"]); res_adm=cap(s["resADM"]); origem="Complemento"
        if not num_jud and s["numJUD"]:
            num_jud=s["numJUD"]; data_jud=fdata(s["dataJUD"]); res_jud=cap(s["resJUD"])
            origem=origem or "Complemento"

    tem_algo = any([data_adm,num_adm,res_adm,data_jud,num_jud,res_jud])
    if origem=="Complemento": n_complemento+=1

    # Regra 2 (sem informação) e Regra 3 (validação mínima: data ADM + nº ADM)
    if not tem_algo:
        validacao = "Sem informação"
        sub = "Ausente no Pipefy/complemento" if not tem_card else "Card sem protocolo (fase inicial)"
    elif data_adm and num_adm:
        validacao = "OK"; sub = ""
    else:
        validacao = "Sem protocolo ADM"; sub = ""

    registros.append({
        "Nome":nome,"CPF":fcpf(cpf),
        "Data Protocolo ADM":data_adm,"Nº Protocolo ADM":num_adm,"Resultado ADM":res_adm,
        "Data Protocolo JUD":data_jud,"Nº Protocolo JUD":num_jud,"Resultado JUD":res_jud,
        "Origem":origem or "—","Validação":validacao,"_sub":sub,"_tem_card":tem_card,
    })
print(f"Preenchidos via complemento: {n_complemento}")

sem_info=[r for r in registros if r["Validação"]=="Sem informação"]
falta_adm=[r for r in registros if r["Validação"]=="Sem protocolo ADM"]
ok=[r for r in registros if r["Validação"]=="OK"]
print(f"Com dados no Pipefy: {n_pipefy} | sem card: {len(registros)-n_pipefy}")
print(f"OK (data+nº ADM): {len(ok)} | Sem protocolo ADM: {len(falta_adm)} | Sem informação: {len(sem_info)}")

# ── Excel ──
wbx=openpyxl.Workbook(); wsx=wbx.active; wsx.title="Processos Waimea"
wsx.merge_cells("A1:J1")
wsx["A1"]=f"FUNDO WAIMEA — Processos (ADM + JUD)  |  {len(registros)} processos"
wsx["A1"].font=Font(bold=True,color="FFFFFF",size=13,name="Calibri")
wsx["A1"].fill=PatternFill("solid",fgColor="0D0E2A"); wsx["A1"].alignment=Alignment(horizontal="center",vertical="center")
wsx.row_dimensions[1].height=26

headers=["Nome","CPF","Data Protocolo ADM","Nº Protocolo ADM","Resultado ADM",
         "Data Protocolo JUD","Nº Protocolo JUD","Resultado JUD","Origem","Validação"]
wsx.append(headers)
font=Font(bold=True,color="FFFFFF"); thin=Side(style="thin",color="BFBFBF")
cf={1:"1F4E78",2:"1F4E78",3:"548235",4:"548235",5:"548235",6:"9E480E",7:"9E480E",8:"9E480E",9:"475569",10:"404040"}
for i,h in enumerate(headers,1):
    c=wsx.cell(2,i); c.fill=PatternFill("solid",fgColor=cf[i]); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=thin)
res_color={"Deferido":"166534","Procedente":"166534","Indeferido":"9A3412","Improcedente":"9A3412","Desistência":"6B7280"}
val_bg={"OK":"DCFCE7","Sem protocolo ADM":"FEF3C7","Sem informação":"FEE2E2"}
for r in registros:
    wsx.append([r[h] for h in headers])
    rr=wsx.max_row
    for ci in (3,4,6,7): wsx.cell(rr,ci).font=Font(name="Courier New",size=9,color="334155"); wsx.cell(rr,ci).alignment=Alignment(horizontal="center")
    for ci,key in ((5,"Resultado ADM"),(8,"Resultado JUD")):
        col=res_color.get(r[key])
        if col: wsx.cell(rr,ci).font=Font(bold=True,color=col,size=9)
    if r["Origem"]=="Complemento": wsx.cell(rr,9).font=Font(bold=True,color="7C3AED",size=9)
    wsx.cell(rr,9).alignment=Alignment(horizontal="center")
    vb=val_bg.get(r["Validação"])
    if vb: wsx.cell(rr,10).fill=PatternFill("solid",fgColor=vb)
    wsx.cell(rr,10).alignment=Alignment(horizontal="center")
wsx.freeze_panes="A3"; wsx.auto_filter.ref=f"A2:J{len(registros)+2}"
for i,w in enumerate([34,15,18,20,16,18,20,16,14,18],1):
    wsx.column_dimensions[get_column_letter(i)].width=w

# ── Aba: Sem Informação (para buscar em outra plataforma) ──
ws3=wbx.create_sheet("Sem Informação")
ws3.append([f"Processos SEM informação no Pipefy — {len(sem_info)} casos (buscar em outra plataforma)"])
ws3["A1"].font=Font(bold=True,color="FFFFFF",size=12); ws3["A1"].fill=PatternFill("solid",fgColor="9A3412")
ws3.merge_cells("A1:C1"); ws3.row_dimensions[1].height=22
ws3.append(["Nome","CPF","Motivo"])
for c in ws3[2]: c.font=font; c.fill=PatternFill("solid",fgColor="1F4E78")
for r in sem_info:
    ws3.append([r["Nome"],r["CPF"],r["_sub"]])
ws3.freeze_panes="A3"
ws3.column_dimensions["A"].width=36; ws3.column_dimensions["B"].width=16; ws3.column_dimensions["C"].width=34

# ── Resumo ──
ws2=wbx.create_sheet("Resumo")
ws2.append(["Métrica","Valor"])
for c in ws2[1]: c.font=font; c.fill=PatternFill("solid",fgColor="1F4E78")
ws2.append(["Total processos",len(registros)])
ws2.append(["Com dados do Pipefy",n_pipefy])
ws2.append(["Preenchidos via complemento",n_complemento])
ws2.append(["Validação mínima OK (data + nº ADM)",len(ok)])
ws2.append(["Sinalizados — sem protocolo ADM",len(falta_adm)])
ws2.append(["Sem informação (à parte)",len(sem_info)])
ws2.append(["  • ausentes no Pipefy",sum(1 for r in sem_info if r['_sub']=='Ausente no Pipefy')])
ws2.append(["  • card em fase inicial s/ protocolo",sum(1 for r in sem_info if r['_sub']!='Ausente no Pipefy')])
ws2.append(["Com protocolo JUD",sum(1 for r in registros if r['Nº Protocolo JUD'])])
ws2.append([])
ws2.append(["Resultado ADM",""])
for k,v in Counter(r["Resultado ADM"] or "(em andamento/sem)" for r in registros).most_common(): ws2.append([k,v])
ws2.append([])
ws2.append(["Resultado JUD",""])
for k,v in Counter(r["Resultado JUD"] or "(sem/—)" for r in registros).most_common(): ws2.append([k,v])
ws2.column_dimensions["A"].width=40; ws2.column_dimensions["B"].width=12

try:
    wbx.save(SAIDA)
    print("Salvo:",os.path.abspath(SAIDA))
except PermissionError:
    alt=SAIDA.replace(".xlsx","_v2.xlsx")
    wbx.save(alt)
    print(f"[!] {os.path.basename(SAIDA)} estava aberto/travado — salvo como:",os.path.abspath(alt))
