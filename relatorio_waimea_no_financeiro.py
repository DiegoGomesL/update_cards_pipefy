# -*- coding: utf-8 -*-
"""
Relatório: processos do FUNDO WAIMEA que estão no pipe FINANCEIRO do Pipefy.

Critério ("estar no financeiro"): o CPF do beneficiário da carteira Waimea
(WAIMEA_900_Processos.xlsx) possui um card no pipe FINANCEIRO (id 305859195).

Para cada match, traz a FASE ATUAL no financeiro + campos financeiros do card,
cruzados com os dados de carteira (cedente, lote, valores).

Saída: ../WAIMEA_no_financeiro.xlsx (na raiz do projeto)
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from lib import api

PIPE_FIN = api.PIPES["FIN"]["id"]
CARTEIRA = os.path.join(os.path.dirname(__file__), "..", "WAIMEA_900_Processos.xlsx")
SAIDA    = os.path.join(os.path.dirname(__file__), "..", "WAIMEA_no_financeiro.xlsx")

# Campos do card FIN que queremos extrair: {field_id: rótulo de saída}
FIN_FIELDS = {
    "fundo_investidor":                        "Fundo Investidor",
    "copy_of_fundo_investidor":                "Fundo Especial",
    "lote_n":                                  "Lote (FIN)",
    "status_da_reuni_o":                        "Status Reunião",
    "valor_do_contrato":                       "Valor Contrato (FIN)",
    "valor_da_entrada":                        "Valor Entrada/Parcela Única",
    "valor_das_parcelas":                      "Valor das Parcelas",
    "valor_para_execu_o":                       "Valor a Executar",
    "valor_do_rpv":                            "Valor RPV",
    "retroativo":                              "Valor Retroativo",
    "data_de_vencimento_da_parcela_em_aberto": "Vencimento Parcela em Aberto",
}

def so_digitos(v):
    return re.sub(r"\D", "", str(v or ""))

def cpf11(v):
    d = so_digitos(v)
    return d.zfill(11) if d and len(d) <= 11 else d

# ── 1) Carteira Waimea ────────────────────────────────────────────────────────
print("Lendo carteira Waimea...")
wb = openpyxl.load_workbook(CARTEIRA, read_only=True, data_only=True)
ws = wb["Processos"]
rows = list(ws.iter_rows(values_only=True))
wb.close()

# header em rows[0]: índices conforme gerar_waimea_completo.py
carteira = {}  # cpf -> dados
for r in rows[1:]:
    cpf = cpf11(r[9])  # 'CPF (11 díg.)'
    if not cpf or len(cpf) != 11:
        continue
    carteira[cpf] = {
        "cedente":         r[1],
        "lote":            r[3],
        "formato":         r[4],
        "processo":        r[6],
        "nome":            r[7],
        "valor_nominal":   r[10],
        "valor_face":      r[11],
        "preco_aquisicao": r[12],
        "vencimento":      r[14],
    }
print(f"  CPFs Waimea na carteira: {len(carteira)}")

# ── 2) Cards do pipe FINANCEIRO ───────────────────────────────────────────────
QUERY = """
query($pipeId: ID!, $after: String) {
  allCards(pipeId: $pipeId, first: 50, after: $after) {
    pageInfo { hasNextPage endCursor }
    edges { node {
      id
      title
      current_phase { name }
      fields { name value field { id } }
    } }
  }
}
"""
print("Buscando cards no pipe FINANCEIRO...")
fin_cards = []  # cada um: {cpf, fase, fields{...}, card_id, title}
after = None
total = 0
while True:
    res = api.execute(QUERY, {"pipeId": PIPE_FIN, "after": after})
    if "errors" in res:
        raise RuntimeError(res["errors"])
    data = res["data"]["allCards"]
    for e in data["edges"]:
        n = e["node"]
        total += 1
        fmap = {(f.get("field") or {}).get("id"): f.get("value") for f in (n.get("fields") or [])}
        fin_cards.append({
            "card_id": n["id"],
            "title":   n.get("title"),
            "fase":    ((n.get("current_phase") or {}).get("name") or "").strip(),
            "cpf":     cpf11(fmap.get("cpf_do_benefici_rio")),
            "nome":    fmap.get("nome_do_benefici_rio"),
            "fmap":    fmap,
        })
    if not data["pageInfo"]["hasNextPage"]:
        break
    after = data["pageInfo"]["endCursor"]
print(f"  Cards no FINANCEIRO: {total}")

# ── 3) Cruzamento: Waimea ∩ FINANCEIRO ────────────────────────────────────────
registros = []
for c in fin_cards:
    info = carteira.get(c["cpf"])
    if not info:
        continue  # card não pertence ao fundo Waimea
    linha = {
        "CPF":               c["cpf"],
        "Nome":              info["nome"] or c["nome"],
        "Cedente":           info["cedente"],
        "Lote":              info["lote"],
        "Formato":           info["formato"],
        "Nº Processo":       info["processo"],
        "Fase Financeiro":   c["fase"],
        "Card ID":           c["card_id"],
        "Valor Nominal":     info["valor_nominal"],
        "Valor Face":        info["valor_face"],
        "Preço Aquisição":   info["preco_aquisicao"],
    }
    for fid, rot in FIN_FIELDS.items():
        linha[rot] = c["fmap"].get(fid)
    registros.append(linha)

# ordena por fase e nome
registros.sort(key=lambda x: (str(x["Fase Financeiro"]), str(x["Nome"])))
print(f"  Processos Waimea NO FINANCEIRO: {len(registros)}")

# ── 4) Excel ──────────────────────────────────────────────────────────────────
out_wb = openpyxl.Workbook()

# Aba resumo por fase
ws_r = out_wb.active
ws_r.title = "Resumo por Fase"
from collections import Counter
fases = Counter(str(r["Fase Financeiro"]) for r in registros)
ws_r.append(["Fase no Financeiro", "Qtd Processos"])
for fase, qtd in sorted(fases.items(), key=lambda x: -x[1]):
    ws_r.append([fase, qtd])
ws_r.append(["TOTAL", len(registros)])

# Aba detalhe
ws_d = out_wb.create_sheet("Processos no Financeiro")
if registros:
    headers = list(registros[0].keys())
else:
    headers = ["CPF", "Nome", "Fase Financeiro"]
ws_d.append(headers)
for r in registros:
    ws_d.append([r.get(h) for h in headers])

# estilo header
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")
for ws_x in (ws_r, ws_d):
    for cell in ws_x[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_x.freeze_panes = "A2"
# larguras razoáveis na aba detalhe
widths = {"CPF":14,"Nome":36,"Cedente":16,"Lote":18,"Formato":12,"Nº Processo":24,
          "Fase Financeiro":26,"Card ID":12}
for i, h in enumerate(headers, 1):
    ws_d.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(h, 20)

out_wb.save(SAIDA)
print(f"\nSalvo: {os.path.abspath(SAIDA)}")
print(f"Total de processos Waimea no financeiro: {len(registros)}")
