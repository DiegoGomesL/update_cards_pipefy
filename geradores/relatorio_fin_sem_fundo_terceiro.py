# -*- coding: utf-8 -*-
"""
Relatório: cards no pipe FINANCEIRO que NÃO têm fundo NEM terceiro interessado.

Critério: FUNDO INVESTIDOR vazio  E  FUNDO ESPECIAL vazio  E  TERCEIRO INTERESSADO vazio.
(São os processos no financeiro ainda não vinculados a nenhum fundo ou terceiro.)

Saída: ../FINANCEIRO_sem_fundo_terceiro.xlsx
"""
import sys, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
import _paths
from lib import api

PIPE_FIN = api.PIPES["FIN"]["id"]
SAIDA = _paths.out("processos","FINANCEIRO_sem_fundo_terceiro.xlsx")

FIN_FIELDS = {
    "lote_n":                                  "Lote",
    "status_da_reuni_o":                        "Status Reunião",
    "valor_do_contrato":                       "Valor Contrato",
    "valor_da_entrada":                        "Valor Entrada/Parcela Única",
    "valor_das_parcelas":                      "Valor das Parcelas",
    "valor_para_execu_o":                       "Valor a Executar",
    "valor_do_rpv":                            "Valor RPV",
    "data_de_vencimento_da_parcela_em_aberto": "Vencimento Parcela em Aberto",
}

QUERY = """
query($pipeId: ID!, $after: String){
  allCards(pipeId:$pipeId, first:50, after:$after){
    pageInfo{hasNextPage endCursor}
    edges{node{ id title current_phase{name} fields{ value field{id} } }}
  }
}
"""

print("Buscando cards no pipe FINANCEIRO...")
after=None; total=0; registros=[]
while True:
    res=api.execute(QUERY,{"pipeId":PIPE_FIN,"after":after})
    if "errors" in res: raise RuntimeError(res["errors"])
    d=res["data"]["allCards"]
    for e in d["edges"]:
        n=e["node"]; total+=1
        fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
        fundo_inv=(fm.get("fundo_investidor") or "").strip()
        fundo_esp=(fm.get("copy_of_fundo_investidor") or "").strip()
        terceiro =(fm.get("terceiro_interessado") or "").strip()
        if fundo_inv or fundo_esp or terceiro:
            continue  # tem fundo ou terceiro -> fora
        linha={
            "CPF":             (fm.get("cpf_do_benefici_rio") or "").strip(),
            "Nome":            (fm.get("nome_do_benefici_rio") or n.get("title") or "").strip(),
            "Fase Financeiro": ((n.get("current_phase") or {}).get("name") or "").strip(),
            "Card ID":         n["id"],
        }
        for fid,rot in FIN_FIELDS.items():
            linha[rot]=fm.get(fid)
        registros.append(linha)
    if not d["pageInfo"]["hasNextPage"]: break
    after=d["pageInfo"]["endCursor"]

registros.sort(key=lambda x:(str(x["Fase Financeiro"]), str(x["Nome"])))
print(f"  Cards no FINANCEIRO: {total}")
print(f"  SEM fundo e SEM terceiro: {len(registros)}")

# Excel
wb=openpyxl.Workbook()
ws_r=wb.active; ws_r.title="Resumo por Fase"
fases=Counter(str(r["Fase Financeiro"]) for r in registros)
ws_r.append(["Fase no Financeiro","Qtd"])
for fase,q in sorted(fases.items(), key=lambda x:-x[1]):
    ws_r.append([fase,q])
ws_r.append(["TOTAL",len(registros)])

ws_d=wb.create_sheet("Sem Fundo e Sem Terceiro")
headers=list(registros[0].keys()) if registros else ["CPF","Nome","Fase Financeiro"]
ws_d.append(headers)
for r in registros:
    ws_d.append([r.get(h) for h in headers])

hdr_fill=PatternFill("solid",fgColor="1F4E78"); hdr_font=Font(bold=True,color="FFFFFF")
for ws_x in (ws_r,ws_d):
    for cell in ws_x[1]:
        cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal="center",vertical="center")
    ws_x.freeze_panes="A2"
widths={"CPF":14,"Nome":36,"Fase Financeiro":26,"Card ID":12,"Lote":16,"Status Reunião":18}
for i,h in enumerate(headers,1):
    ws_d.column_dimensions[get_column_letter(i)].width=widths.get(h,18)

wb.save(SAIDA)
print(f"\nSalvo: {os.path.abspath(SAIDA)}")
