# -*- coding: utf-8 -*-
"""
Relatório: PROCESSOS LIVRES EM ANDAMENTO (disponíveis para cessão a fundo).

Critérios de elegibilidade (consolidado por CPF, juntando ADM + JUD):
  1. Os 3 campos de vínculo VAZIOS em TODOS os cards (ADM e JUD) do CPF:
     TERCEIRO INTERESSADO, FUNDO INVESTIDOR, FUNDO ESPECIAL.
  2. NÃO estar no financeiro: CPF sem card no pipe FINANCEIRO.
  3. EM ANDAMENTO: ao menos 1 card ativo
     - ADM fora de {Encerrados, Descartados}
     - JUD fora de {Encerrados, Procedentes}
  4. SEM resultado final favorável (sairia para o financeiro):
     - ADM != DEFERIDO  e  JUD != Procedente (processo/recurso/fase Procedentes)

Consolidação: 1 linha por CPF. Escolhe o card ADM e o JUD mais relevantes
(prefere card ativo; senão o de protocolo mais recente). Cobre ADM+JUD
simultâneos e retorno JUD->ADM (novo protocolo).

Saída: ../PROCESSOS_LIVRES_EM_ANDAMENTO.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import _paths
from lib import api

ADM = api.PIPES["ADM"]["id"]
JUD = api.PIPES["JUD"]["id"]
FIN = api.PIPES["FIN"]["id"]
SAIDA = _paths.out("processos","PROCESSOS_LIVRES_EM_ANDAMENTO.xlsx")

# Field IDs por pipe
F = {
    "ADM": {"cpf":"cpf_do_benefici_rio_1", "nome":"nome_do_benefici_rio",
            "terceiro":"terceiro_interessado", "fundo_inv":"fundo_investidor",
            "fundo_esp":"copy_of_fundo_investidor",
            "protocolo":"n_mero_do_processo_administrativo", "data_prot":"data_do_protocolo",
            "resultado":"resultado_do_pedido"},
    "JUD": {"cpf":"cpf_do_benefici_rio", "nome":"nome_do_benefici_rio",
            "terceiro":"terceiro_interassado", "fundo_inv":"fundo_investidor",
            "fundo_esp":"copy_of_fundo_investidor",
            "protocolo":"n_mero_do_processo", "data_prot":"data_do_protocolo",
            "resultado":"resultado_do_processo", "recurso":"resultado_do_recurso"},
}

ADM_TERMINAL = {"ENCERRADOS", "DESCARTADOS"}
JUD_TERMINAL = {"ENCERRADOS", "PROCEDENTES"}

def dig(v): return re.sub(r"\D", "", str(v or ""))
def cpf11(v):
    d = dig(v); return d.zfill(11) if d and len(d) <= 11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fmt_cpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c) == 11 else c
def fmt_data(v):
    s = nz(v)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else s

CARDS_Q = """
query($pipeId: ID!, $after: String){
  allCards(pipeId:$pipeId, first:50, after:$after){
    pageInfo{hasNextPage endCursor}
    edges{node{ id title current_phase{name} fields{ value field{id} } }}
  }
}
"""
def pull(pid, kind):
    out=[]; after=None
    while True:
        r=api.execute(CARDS_Q,{"pipeId":pid,"after":after})
        if "errors" in r: raise RuntimeError(r["errors"])
        d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]
            fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fids=F[kind]
            phase=nz((n.get("current_phase") or {}).get("name"))
            out.append({
                "kind":kind, "card_id":n["id"], "title":nz(n.get("title")), "phase":phase,
                "cpf":cpf11(fm.get(fids["cpf"])), "nome":nz(fm.get(fids["nome"])),
                "terceiro":nz(fm.get(fids["terceiro"])), "fundo_inv":nz(fm.get(fids["fundo_inv"])),
                "fundo_esp":nz(fm.get(fids["fundo_esp"])),
                "protocolo":nz(fm.get(fids["protocolo"])), "data_prot":nz(fm.get(fids["data_prot"])),
                "resultado":nz(fm.get(fids["resultado"])),
                "recurso":nz(fm.get(fids.get("recurso"))) if fids.get("recurso") else "",
            })
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando cards ADM...");  adm_cards = pull(ADM,"ADM")
print(f"  ADM: {len(adm_cards)}")
print("Buscando cards JUD...");  jud_cards = pull(JUD,"JUD")
print(f"  JUD: {len(jud_cards)}")
print("Buscando CPFs no FINANCEIRO...")
fin_cpfs=set()   # FIN usa o field id cpf_do_benefici_rio
after=None
while True:
    r=api.execute(CARDS_Q,{"pipeId":FIN,"after":after}); d=r["data"]["allCards"]
    for e in d["edges"]:
        fm={(f.get("field") or {}).get("id"):f.get("value") for f in (e["node"].get("fields") or [])}
        c=cpf11(fm.get("cpf_do_benefici_rio"))
        if c: fin_cpfs.add(c)
    if not d["pageInfo"]["hasNextPage"]: break
    after=d["pageInfo"]["endCursor"]
print(f"  CPFs no FINANCEIRO: {len(fin_cpfs)}")

# Agrupar por CPF
from collections import defaultdict
adm_by=defaultdict(list); jud_by=defaultdict(list)
for c in adm_cards:
    if c["cpf"]: adm_by[c["cpf"]].append(c)
for c in jud_cards:
    if c["cpf"]: jud_by[c["cpf"]].append(c)

def is_active(c):
    ph=c["phase"].upper()
    term = ADM_TERMINAL if c["kind"]=="ADM" else JUD_TERMINAL
    return ph not in term

def pick(cards):
    if not cards: return None
    ativos=[c for c in cards if is_active(c)]
    pool=ativos if ativos else cards
    return max(pool, key=lambda c: (c["data_prot"], c["protocolo"]))

def ganhou(cpf):
    for c in adm_by.get(cpf,[]):
        if c["resultado"].upper()=="DEFERIDO": return True
    for c in jud_by.get(cpf,[]):
        if "PROCEDENTE" in c["resultado"].upper(): return True
        if c["phase"].upper()=="PROCEDENTES": return True
        if "PROCEDENTE" in c["recurso"].upper(): return True
    return False

def tem_vinculo(cpf):
    for c in adm_by.get(cpf,[])+jud_by.get(cpf,[]):
        if c["terceiro"] or c["fundo_inv"] or c["fundo_esp"]: return True
    return False

registros=[]
todos_cpfs=set(adm_by)|set(jud_by)
for cpf in todos_cpfs:
    if cpf in fin_cpfs:        continue   # já no financeiro
    if tem_vinculo(cpf):       continue   # vinculado a fundo/terceiro
    if ganhou(cpf):            continue   # resultado final favorável
    a=pick(adm_by.get(cpf,[])); j=pick(jud_by.get(cpf,[]))
    a_active = a and is_active(a); j_active = j and is_active(j)
    if not (a_active or j_active): continue  # nenhum card ativo -> não está em andamento

    nome = (a or j)["nome"] or (j or a)["nome"]
    jph = j["phase"] if j else ""
    registros.append({
        "Nome": nome,
        "CPF": fmt_cpf(cpf),
        "Terceiro Interessado": "",
        "Fundo Investidor": "",
        "Fundo Especial": "",
        "Fase do card (ADM)":      a["phase"]     if a else "",
        "Nº Protocolo (ADM)":      a["protocolo"] if a else "",
        "Data Protocolo (ADM)":    fmt_data(a["data_prot"]) if a else "",
        "Resultado Administrativo":a["resultado"] if a else "",
        "Fase do card (JUD)":      jph,
        "Nº Protocolo (JUD)":      j["protocolo"] if j else "",
        "Data Protocolo (JUD)":    fmt_data(j["data_prot"]) if j else "",
        "Fase 8 - Aguardando Julgamento (1ª inst.)": "Sim" if jph.upper().startswith("08") else "",
        "Fase 9 - Em Grau Recursal (2ª inst.)":      "Sim" if jph.upper().startswith("09") else "",
        "Resultado Recurso Inominado": j["recurso"] if j else "",
    })

registros.sort(key=lambda x:(x["Fase do card (JUD)"], x["Fase do card (ADM)"], str(x["Nome"])))
print(f"\nProcessos livres em andamento: {len(registros)}")

# ── Excel ──────────────────────────────────────────────────────────────
wb=openpyxl.Workbook()
ws=wb.active; ws.title="Processos Livres"
headers=list(registros[0].keys()) if registros else ["Nome","CPF"]
ws.append(headers)
for r in registros: ws.append([r.get(h) for h in headers])

# estilo: cabeçalho com grupos de cor (Identificação / ADM / JUD)
GRP = {  # índice 1-based -> cor
}
ident = range(1,6); adm_cols = range(6,10); jud_cols = range(10,16)
col_fill = {}
for i in ident:    col_fill[i]="1F4E78"  # azul
for i in adm_cols: col_fill[i]="548235"  # verde
for i in jud_cols: col_fill[i]="9E480E"  # laranja
font=Font(bold=True,color="FFFFFF")
thin=Side(style="thin",color="BFBFBF")
for i,h in enumerate(headers,1):
    cell=ws.cell(row=1,column=i)
    cell.fill=PatternFill("solid",fgColor=col_fill.get(i,"404040"))
    cell.font=font
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=Border(bottom=thin)
ws.freeze_panes="A2"; ws.row_dimensions[1].height=42
widths=[34,15,18,16,15, 26,22,16,24, 26,22,16, 18,18,24]
for i,w in enumerate(widths[:len(headers)],1):
    ws.column_dimensions[get_column_letter(i)].width=w

# aba resumo
ws2=wb.create_sheet("Resumo")
from collections import Counter
ws2.append(["Métrica","Valor"])
ws2.append(["Total processos livres em andamento", len(registros)])
ws2.append(["Com protocolo ADM", sum(1 for r in registros if r["Nº Protocolo (ADM)"])])
ws2.append(["Com protocolo JUD", sum(1 for r in registros if r["Nº Protocolo (JUD)"])])
ws2.append(["Sem nenhum protocolo", sum(1 for r in registros if not r["Nº Protocolo (ADM)"] and not r["Nº Protocolo (JUD)"])])
ws2.append(["Em Fase 8 (Aguardando Julgamento)", sum(1 for r in registros if r["Fase 8 - Aguardando Julgamento (1ª inst.)"]=="Sim")])
ws2.append(["Em Fase 9 (Grau Recursal)", sum(1 for r in registros if r["Fase 9 - Em Grau Recursal (2ª inst.)"]=="Sim")])
ws2.append([])
ws2.append(["Distribuição por Fase ADM",""])
for k,v in Counter(r["Fase do card (ADM)"] or "(sem ADM)" for r in registros).most_common():
    ws2.append([k,v])
ws2.append([])
ws2.append(["Distribuição por Fase JUD",""])
for k,v in Counter(r["Fase do card (JUD)"] or "(sem JUD)" for r in registros).most_common():
    ws2.append([k,v])
for cell in ws2[1]:
    cell.fill=PatternFill("solid",fgColor="1F4E78"); cell.font=font
ws2.column_dimensions["A"].width=42; ws2.column_dimensions["B"].width=14

wb.save(SAIDA)
print(f"Salvo: {os.path.abspath(SAIDA)}")
