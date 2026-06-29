# -*- coding: utf-8 -*-
"""Consolida os números do Fundo ENIO para a apresentação -> ENIO_Dados_Apresentacao.xlsx"""
import json
import _paths
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

fin = json.load(open(_paths.dados('enio_financeiro.json'), encoding='utf-8'))
met = json.load(open(_paths.dados('enio_metrics.json'), encoding='utf-8'))

CONTRATO   = 412500.00
INVEST     = 150000.00
num = lambda x: float(x or 0)

ganho_adm_n = met['ganhos_adm_deferido']           # 26
ganho_jud_n = met['ganhos_jud_procedente']         # 15
ganhos_n    = ganho_adm_n + ganho_jud_n            # 41

adm = fin['deferido_adm']; jud = fin['procedente_jud']; todos = fin['todos']
adm_total = num(adm['total']);  adm_fin = int(adm['cpfs'])
jud_total = num(jud['total']);  jud_fin = int(jud['cpfs'])
# valor/proc = total recuperável ÷ nº de ganhos (consistente: Proc × Valor/proc = Total)
adm_vproc = adm_total/ganho_adm_n
jud_vproc = jud_total/ganho_jud_n
lastro_total = adm_total + jud_total

recebido  = num(todos['recebido'])
a_receber = num(todos['em_atraso']) + num(todos['no_prazo']) + num(todos['sem_info'])
saldo     = CONTRATO - recebido
pct_receb = recebido / CONTRATO
pct_saldo = saldo / CONTRATO
cobertura_map = (recebido + a_receber) / CONTRATO

taxa_adm = met['taxa_adm']; taxa_jud = met['taxa_jud']
pipe_adm = met['pipeline_adm_aguardando']; pipe_jud = met['pipeline_jud_aguardando']
proj_adm = round(pipe_adm * taxa_adm); proj_jud = round(pipe_jud * taxa_jud)
proj_n   = proj_adm + proj_jud
val_adm  = proj_adm * adm_vproc; val_jud = proj_jud * jud_vproc
pipe_val = val_adm + val_jud

lastro_total_proj = lastro_total + pipe_val
total_proj_n = ganhos_n + proj_n
cobertura_proj = lastro_total_proj / CONTRATO

# ── Excel ──
NAVY="1E204E"; GOLD="F5A623"; GREEN="166534"; GREENBG="DCFCE7"; HDR="1F4E78"; WHITE="FFFFFF"; ALT="F2F6FA"
wb=Workbook(); ws=wb.active; ws.title="Dados Apresentação"
B=Side(style='thin',color="BFC8D6"); bd=Border(left=B,right=B,top=B,bottom=B)
def put(r,c,v,bold=False,bg=None,col="1E293B",al="left",sz=11,money=False):
    cell=ws.cell(r,c, ('R$ %s'%format(v,',.2f').replace(',','X').replace('.',',').replace('X','.')) if money and isinstance(v,(int,float)) else v)
    cell.font=Font(bold=bold,color=col,size=sz,name="Calibri")
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal=al,vertical='center',wrap_text=True); cell.border=bd
    return cell
def title(r,txt):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    put(r,1,txt,bold=True,bg=NAVY,col=WHITE,sz=12)
    ws.row_dimensions[r].height=22
def hdr(r,cols):
    for i,c in enumerate(cols,1): put(r,i,c,bold=True,bg=HDR,col=WHITE,al='center')
def pct(x): return format(x*100,'.1f').replace('.',',')+'%'

ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=16
ws.column_dimensions['C'].width=18; ws.column_dimensions['D'].width=18

r=1
ws.merge_cells('A1:D1'); put(1,1,"FUNDO ENIO — Dados para Apresentação (valores reais BD + status Pipefy)",bold=True,bg=NAVY,col=WHITE,sz=13)
ws.row_dimensions[1].height=26
r=3
title(r,"CONTRATO"); r+=1
hdr(r,["Item","Valor","",""]); r+=1
put(r,1,"Valor do contrato (BD)"); put(r,2,CONTRATO,money=True,al='right'); r+=1
put(r,1,"Investimento do fundo"); put(r,2,INVEST,money=True,al='right'); r+=1
put(r,1,"Processos contratados"); put(r,2,129,al='right'); r+=2

title(r,"1 — GANHOS REALIZADOS (LASTRO)"); r+=1
hdr(r,["Via","Proc.","Valor/Proc. (BD)","Total recuperável"]); r+=1
put(r,1,"ADM — deferidos"); put(r,2,ganho_adm_n,al='center'); put(r,3,adm_vproc,money=True,al='right'); put(r,4,adm_total,money=True,al='right'); r+=1
put(r,1,"JUD — procedentes"); put(r,2,ganho_jud_n,al='center'); put(r,3,jud_vproc,money=True,al='right'); put(r,4,jud_total,money=True,al='right'); r+=1
put(r,1,"Total",bold=True,bg=GREENBG); put(r,2,ganhos_n,bold=True,bg=GREENBG,al='center'); put(r,3,"—",bold=True,bg=GREENBG,al='right'); put(r,4,lastro_total,bold=True,bg=GREENBG,col=GREEN,money=True,al='right'); r+=2

title(r,"DESEMPENHO DO CONTRATO"); r+=1
hdr(r,["Indicador","Valor","",""]); r+=1
put(r,1,"Valor do contrato (ENIO)"); put(r,2,CONTRATO,money=True,al='right'); r+=1
put(r,1,"Já recebido (amortizado)"); put(r,2,f"R$ {recebido:,.2f}  ({pct(pct_receb)})".replace(',','X').replace('.',',').replace('X','.'),col=GREEN,al='right'); r+=1
put(r,1,"Saldo a amortizar"); put(r,2,f"R$ {saldo:,.2f}  ({pct(pct_saldo)})".replace(',','X').replace('.',',').replace('X','.'),al='right'); r+=1
put(r,1,"A receber na esteira"); put(r,2,a_receber,money=True,al='right'); r+=1
put(r,1,"Cobertura mapeada (receb.+esteira)",bold=True,bg=GREENBG); put(r,2,pct(cobertura_map),bold=True,bg=GREENBG,col=GREEN,al='right'); r+=2

title(r,"2 — PROJEÇÃO (pipeline)"); r+=1
hdr(r,["Etapa","Proc.","Taxa hist.","Proj. ganhos"]); r+=1
put(r,1,"Aguardando Decisão ADM"); put(r,2,pipe_adm,al='center'); put(r,3,pct(taxa_adm),al='center'); put(r,4,f"~{proj_adm}",al='center'); r+=1
put(r,1,"Aguardando Julgamento JUD"); put(r,2,pipe_jud,al='center'); put(r,3,pct(taxa_jud),al='center'); put(r,4,f"~{proj_jud}",al='center'); r+=1
put(r,1,"Total pipeline",bold=True,bg=ALT); put(r,2,pipe_adm+pipe_jud,bold=True,bg=ALT,al='center'); put(r,3,"—",bold=True,bg=ALT,al='center'); put(r,4,f"~{proj_n}",bold=True,bg=ALT,al='center'); r+=1
put(r,1,"Valor estimado pipeline",bold=True); put(r,2,"",);
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
put(r,2,f"ADM ~{proj_adm}×R$ {adm_vproc:,.0f} + JUD ~{proj_jud}×R$ {jud_vproc:,.0f} = R$ {pipe_val:,.2f}".replace(',','.'),al='right'); r+=2

title(r,"PROJEÇÃO CONSOLIDADA"); r+=1
hdr(r,["Item","Processos","Valor (lastro)",""]); r+=1
put(r,1,"Lastro realizado (deferido/procedente)"); put(r,2,ganhos_n,al='center'); put(r,3,lastro_total,money=True,al='right'); r+=1
put(r,1,"Pipeline projetado"); put(r,2,f"+{proj_n}",al='center'); put(r,3,pipe_val,money=True,al='right'); r+=1
put(r,1,"Lastro total projetado",bold=True,bg=GREENBG); put(r,2,total_proj_n,bold=True,bg=GREENBG,al='center'); put(r,3,lastro_total_proj,bold=True,bg=GREENBG,col=GREEN,money=True,al='right'); r+=1
put(r,1,"Cobertura do contrato (R$ 412,5K)",bold=True); put(r,2,"—",al='center'); put(r,3,pct(cobertura_proj),bold=True,col=GREEN,al='right'); r+=2

title(r,"SITUAÇÃO DA CARTEIRA (209 processos)"); r+=1
hdr(r,["Situação","Qtd","",""]); r+=1
for k in ["FIN","ADM","JUD","ENCERRADO"]:
    put(r,1,{"FIN":"FIN — Ganhos (ADM+JUD)","ADM":"ADM — Aguardando decisão","JUD":"JUD — Aguardando sentença","ENCERRADO":"ENCERRADO — Perdas/Desist."}[k])
    put(r,2,met['por_situacao'].get(k,0),al='center'); r+=1
put(r,1,"TOTAL",bold=True,bg=ALT); put(r,2,met['total'],bold=True,bg=ALT,al='center'); r+=1

out=_paths.out('enio','ENIO_Dados_Apresentacao.xlsx')
wb.save(out)
print("Salvo:",out)
# imprimir resumo no console
print(f"""
CONTRATO            R$ {CONTRATO:,.2f}  | investimento R$ {INVEST:,.2f} | 129 processos
GANHOS (lastro)     {ganhos_n}  (ADM {ganho_adm_n} + JUD {ganho_jud_n})  = R$ {lastro_total:,.2f}
  ADM deferidos     {ganho_adm_n} | R$/proc {adm_vproc:,.2f} | total R$ {adm_total:,.2f}
  JUD procedentes   {ganho_jud_n} | R$/proc {jud_vproc:,.2f} | total R$ {jud_total:,.2f}
DESEMPENHO
  Já recebido       R$ {recebido:,.2f} ({pct(pct_receb)})
  Saldo             R$ {saldo:,.2f} ({pct(pct_saldo)})
  A receber esteira R$ {a_receber:,.2f}
  Cobertura mapeada {pct(cobertura_map)}
PROJEÇÃO
  ADM {pipe_adm} x {pct(taxa_adm)} = ~{proj_adm} | JUD {pipe_jud} x {pct(taxa_jud)} = ~{proj_jud} | total ~{proj_n}
  valor pipeline    R$ {pipe_val:,.2f}
CONSOLIDADA
  total projetado   {total_proj_n} proc | R$ {lastro_total_proj:,.2f} | cobertura {pct(cobertura_proj)}
""")
