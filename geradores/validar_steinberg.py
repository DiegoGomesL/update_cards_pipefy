# -*- coding: utf-8 -*-
"""Valida a aba 'Steinberg' do RELATORIO_UNIFICADO — os processos estão livres no Pipefy?
Livre = 3 vínculos vazios (Terceiro/Fundo Inv/Fundo Esp) em ADM+JUD+FIN + NÃO no financeiro + em andamento.
Casamento por CPF. Saída: relatorios/validacoes/Steinberg_VALIDADO.xlsx"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
SRC=r'C:\Users\DIEGO T&C\Downloads\RELATORIO_UNIFICADO_TelesCosta_Steinberg.xlsx'
SAIDA=_paths.out("validacoes","Steinberg_VALIDADO.xlsx")

FID={"ADM":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio_1","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "JUD":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interassado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "FIN":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS","PROCEDENTES"}
def dig(v): return re.sub(r"\D","",str(v or ""))
def c11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=FID[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":c11(fm.get(fi["cpf"])),"terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
by=defaultdict(lambda:{"ADM":[],"JUD":[],"FIN":[]})
for c in pull(ADM,"ADM")+pull(JUD,"JUD")+pull(FIN,"FIN"):
    if c["cpf"]: by[c["cpf"]][c["kind"]].append(c)

wb=openpyxl.load_workbook(SRC,data_only=True); ws=wb["Steinberg"]
casos=[(nz(r[1]),c11(r[2])) for r in ws.iter_rows(min_row=2,values_only=True) if r[2]]
print(f"Casos Steinberg: {len(casos)}\n")

def active(c):
    term=ADM_TERM if c["kind"]=="ADM" else (JUD_TERM if c["kind"]=="JUD" else set())
    return c["phase"].upper() not in term

regs=[]; livre=0; naolivre=0; naoenc=0
for nome,cpf in casos:
    d=by.get(cpf)
    if not d:
        regs.append([nome,fcpf(cpf),"NÃO ENCONTRADO","CPF sem card no Pipefy (ainda não cadastrado)","","",""]); naoenc+=1; continue
    todos=d["ADM"]+d["JUD"]+d["FIN"]
    tadm=next((c["terc"] for c in d["ADM"] if c["terc"]),"")
    tjud=next((c["terc"] for c in d["JUD"] if c["terc"]),"")
    tfin=next((c["terc"] for c in d["FIN"] if c["terc"]),"")
    inv=next((c["inv"] for c in todos if c["inv"]),"")
    esp=next((c["esp"] for c in todos if c["esp"]),"")
    in_fin=bool(d["FIN"]); tem_ativo=any(active(c) for c in d["ADM"]+d["JUD"])
    mot=[]; vinc=[]
    if tadm: vinc.append(f"Terc.ADM={tadm}")
    if tjud: vinc.append(f"Terc.JUD={tjud}")
    if tfin: vinc.append(f"Terc.FIN={tfin}")
    if inv: vinc.append(f"FundoInv={inv}")
    if esp: vinc.append(f"FundoEsp={esp}")
    if vinc: mot.append("VÍNCULO: "+"; ".join(vinc))
    if in_fin: mot.append("NO FINANCEIRO")
    if not tem_ativo: mot.append("sem card ativo")
    ver="LIVRE" if not mot else "NÃO LIVRE"
    if not mot: livre+=1
    else: naolivre+=1
    fadm=next((c["phase"] for c in d["ADM"]),""); fjud=next((c["phase"] for c in d["JUD"]),"")
    regs.append([nome,fcpf(cpf),ver," | ".join(mot),fadm,fjud,"Sim" if in_fin else ""])

# Excel
out=openpyxl.Workbook(); wsx=out.active; wsx.title="Validação Steinberg"
headers=["Nome (Steinberg)","CPF","Veredito","Motivo","Fase ADM","Fase JUD","No Financeiro?"]
wsx.append(headers)
GREEN="DCFCE7"; RED="FEE2E2"; GRAY="F1F5F9"
font=Font(bold=True,color="FFFFFF")
for i,h in enumerate(headers,1):
    c=wsx.cell(1,i); c.fill=PatternFill("solid",fgColor="1F4E78"); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=Side(style="thin",color="BFBFBF"))
for r in regs:
    wsx.append(r); rr=wsx.max_row
    bg=GREEN if r[2]=="LIVRE" else (GRAY if r[2]=="NÃO ENCONTRADO" else RED)
    wsx.cell(rr,3).fill=PatternFill("solid",fgColor=bg)
wsx.freeze_panes="A2"
for i,w in enumerate([32,15,16,52,24,24,13],1): wsx.column_dimensions[get_column_letter(i)].width=w
try: out.save(SAIDA); print("Salvo:",SAIDA)
except PermissionError:
    alt=SAIDA.replace(".xlsx","_v2.xlsx"); out.save(alt); print("[!] travado — salvo como:",alt)

print(f"\nLIVRES: {livre} | NÃO LIVRES: {naolivre} | NÃO ENCONTRADOS (fora do Pipefy): {naoenc} | total {len(regs)}")
print("\nDetalhe:")
for r in regs: print(f"  {r[2]:15} {r[0][:30]:30} {r[1]}  {r[3][:60]}")
