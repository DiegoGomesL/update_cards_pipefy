import json, re
import _paths
import openpyxl as _opxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _norm_cpf(cpf): return re.sub(r'[^0-9]', '', str(cpf or ''))

# JUD process numbers — lotes ENIO da BASE JUD
_jud_wb = _opxl.load_workbook(_paths.src('BASE JUD.xlsx'))
_jud_ws = _jud_wb.active
jud_map = {}
for _row in _jud_ws.iter_rows(values_only=True, min_row=2):
    _ti = str(_row[5] or '').upper()
    if 'ENIO' in _ti:
        _cpf = _norm_cpf(_row[2]); _proc = _row[7]
        if _cpf and _proc: jud_map[_cpf] = str(_proc).strip()

with open(_paths.dados('enio_casos.json'), encoding='utf-8') as f:
    casos = json.load(f)

def lote_label(raw):
    u = (raw or '').upper()
    if 'FILHO' in u and 'ADITIVO' in u: return 'ENIO Filho Aditivo'
    if 'ADITIVO' in u:                  return 'ENIO Aditivo'
    if 'FILHO' in u:                    return 'ENIO Filho 23/01/25'
    if '27/11' in u or '27.11' in u:    return 'ENIO 27/11/25'
    return 'ENIO'

LOTE_COLOR = {
    'ENIO Filho 23/01/25': ("1E204E", "FFFFFF"),
    'ENIO 27/11/25':       ("7C3AED", "FFFFFF"),
    'ENIO Filho Aditivo':  ("0E7490", "FFFFFF"),
    'ENIO Aditivo':        ("B45309", "FFFFFF"),
    'ENIO':                ("475569", "FFFFFF"),
}

NAVY="0D0E2A"; GOLD="F5A623"; GREEN="22C55E"; GRAY="64748B"; WHITE="FFFFFF"; L_BLUE="EEF2FF"
def fill(h): return PatternFill("solid", fgColor=h)
def font(h, sz=10, bold=False): return Font(color=h, size=sz, bold=bold, name="Calibri")
def border():
    s=Side(style='thin', color="CBD5E1"); return Border(left=s,right=s,top=s,bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)

for c in casos: c['lote_label'] = lote_label(c.get('lote'))

fin_count=sum(1 for c in casos if c['situacao']=='FIN')
adm_count=sum(1 for c in casos if c['situacao']=='ADM')
jud_count=sum(1 for c in casos if c['situacao']=='JUD')
enc_count=sum(1 for c in casos if c['situacao']=='ENCERRADO')

wb=Workbook(); ws=wb.active; ws.title="ENIO — Casos"

ws.merge_cells("A1:G1")
ws["A1"]="FUNDO ENIO — Lista Completa de Casos Previdenciários (Todos os Lotes)"
ws["A1"].font=Font(color=WHITE,size=14,bold=True,name="Calibri"); ws["A1"].fill=fill(NAVY); ws["A1"].alignment=center()
ws.row_dimensions[1].height=28

ws.merge_cells("A2:G2")
ws["A2"]=(f"Total: {len(casos)} processos  |  FIN: {fin_count}  |  ADM: {adm_count}  |  "
          f"JUD: {jud_count}  |  Encerrado: {enc_count}  |  Teles & Costa")
ws["A2"].font=Font(color=GOLD,size=9,bold=True,name="Calibri"); ws["A2"].fill=fill("161838"); ws["A2"].alignment=center()
ws.row_dimensions[2].height=16

headers=["Nome do Beneficiário","CPF","Lote","Nº Processo ADM","Nº Processo JUD","Situação","Resultado"]
col_widths=[36,16,20,18,18,13,30]
for i,(h,w) in enumerate(zip(headers,col_widths),start=1):
    cell=ws.cell(row=3,column=i,value=h)
    cell.font=font(WHITE,sz=10,bold=True); cell.fill=fill("1E204E"); cell.alignment=center(); cell.border=border()
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[3].height=22

SIT_STYLE={"FIN":(GREEN,WHITE),"ADM":(GOLD,NAVY),"JUD":("3B82F6",WHITE),"ENCERRADO":(GRAY,WHITE)}
RES_LABEL={"DEFERIDO":"Deferido","INDEFERIDO":"Indeferido","PROCEDENTE":"Procedente (JUD)",
    "IMPROCEDENTE":"Improcedente (JUD)","DESISTÊNCIA":"Desistência","DESISTENCIA":"Desistência",
    "EXTINÇÃO SEM ANÁLISE DE MÉRITO":"Extinção s/ Mérito","EM ANDAMENTO":"Em Andamento"}
def res_label(r):
    if not r: return "Em Andamento"
    u=r.upper()
    for k,v in RES_LABEL.items():
        if u==k or u.startswith(k) or k in u: return v
    return r.title()

for idx,caso in enumerate(casos):
    row=idx+4; is_even=idx%2==0
    sit=caso.get("situacao","ADM"); lote=caso.get("lote_label","ENIO")
    sit_color,sit_text=SIT_STYLE.get(sit,(GRAY,WHITE))
    lote_bg,lote_fg=LOTE_COLOR.get(lote,("475569",WHITE))
    resultado=res_label(caso.get("resultado_final") or "EM ANDAMENTO")
    row_bg = "FFF7ED" if sit=="ENCERRADO" else ("F0FDF4" if sit=="FIN" else (WHITE if is_even else L_BLUE))
    values=[caso.get("nome",""), caso.get("cpf",""), lote,
            caso.get("num_proc_adm","") or "",
            jud_map.get(_norm_cpf(caso.get("cpf","")),""), sit, resultado]
    for col,val in enumerate(values,start=1):
        cell=ws.cell(row=row,column=col,value=val); cell.border=border(); cell.font=font("1E293B",sz=9)
        if col in (4,5): cell.alignment=center(); cell.font=Font(color="334155",size=9,name="Courier New")
        elif col==2:     cell.alignment=center()
        else:            cell.alignment=left()
        cell.fill=fill(row_bg)
        if col==3:   cell.fill=fill(lote_bg); cell.font=font(lote_fg,sz=8,bold=True); cell.alignment=center()
        elif col==6: cell.fill=fill(sit_color); cell.font=font(sit_text,sz=9,bold=True); cell.alignment=center()
    ws.row_dimensions[row].height=16

ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:G{3+len(casos)}"

# ── Resumo ─────────────────────────────────────────────
ws2=wb.create_sheet("Resumo")
ws2.merge_cells("A1:C1")
ws2["A1"]="ENIO — Resumo por Situação"
ws2["A1"].font=Font(color=WHITE,size=12,bold=True,name="Calibri"); ws2["A1"].fill=fill(NAVY); ws2["A1"].alignment=center()
ws2.row_dimensions[1].height=24
resumo=[("Situação","Qtd",""),
    ("FIN — Ganhos (ADM + JUD)",fin_count,GREEN),
    ("ADM — Aguardando decisão",adm_count,GOLD),
    ("JUD — Aguardando sentença",jud_count,"3B82F6"),
    ("ENCERRADO — Perdas/Desist.",enc_count,GRAY),
    ("TOTAL",len(casos),NAVY)]
for r,(label,qtd,color) in enumerate(resumo,start=2):
    is_hdr=r==2
    ws2.cell(r,1,label).font=font(GOLD if is_hdr else WHITE,sz=10,bold=is_hdr)
    ws2.cell(r,1).fill=fill("1E204E" if is_hdr else color); ws2.cell(r,1).alignment=left(); ws2.cell(r,1).border=border()
    ws2.cell(r,2,"" if is_hdr else qtd).font=font(WHITE,sz=10,bold=(r==len(resumo)+1))
    ws2.cell(r,2).fill=fill("1E204E" if is_hdr else color); ws2.cell(r,2).alignment=center(); ws2.cell(r,2).border=border()
    ws2.row_dimensions[r].height=20

ws2.merge_cells("A9:C9")
ws2["A9"]="Por Lote"; ws2["A9"].font=Font(color=WHITE,size=11,bold=True,name="Calibri")
ws2["A9"].fill=fill(NAVY); ws2["A9"].alignment=center(); ws2.row_dimensions[9].height=22
from collections import Counter
lote_cnt=Counter(c['lote_label'] for c in casos)
ri=10
for lname,cnt in sorted(lote_cnt.items(), key=lambda x:-x[1]):
    lcolor=LOTE_COLOR.get(lname,("475569","FFFFFF"))[0]
    ws2.cell(ri,1,lname).font=font(WHITE,sz=10); ws2.cell(ri,1).fill=fill(lcolor); ws2.cell(ri,1).alignment=left(); ws2.cell(ri,1).border=border()
    ws2.cell(ri,2,cnt).font=font(WHITE,sz=10,bold=True); ws2.cell(ri,2).fill=fill(lcolor); ws2.cell(ri,2).alignment=center(); ws2.cell(ri,2).border=border()
    ws2.row_dimensions[ri].height=20; ri+=1
ws2.column_dimensions["A"].width=34; ws2.column_dimensions["B"].width=10

out=_paths.out('enio','ENIO_Casos_v2.xlsx')
wb.save(out)
print(f"Planilha salva: {out}")
print(f"Total: {len(casos)} | FIN: {fin_count} | ADM: {adm_count} | JUD: {jud_count} | ENCERRADO: {enc_count}")
print(f"Com Nº JUD: {sum(1 for c in casos if jud_map.get(_norm_cpf(c.get('cpf',''))))}")
print("Lotes:", dict(lote_cnt))
