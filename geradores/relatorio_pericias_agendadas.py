# -*- coding: utf-8 -*-
"""Relatório de perícias médicas AGENDADAS para os próximos meses (Ago–Dez 2026).

Contexto: o INSS vem antecipando perícias sem aviso prévio. Este relatório lista
todas as perícias agendadas mês a mês (uma aba por mês) para que a equipe divida a
carga de verificação individual de cada caso.

Fonte: pipe "PERÍCIA MÉDICA ADM" (305779737) do Pipefy — puxado AO VIVO a cada
execução (o espelho no banco está desatualizado). Data considerada = campo oficial
"DATA E HORA DA PERÍCIA MÉDICA (VENC)". Exclui cards na fase CONCLUÍDO (perícia já
encerrada, fora do worklist).

Saída: relatorios/pericias/PERICIAS_AGENDADAS_AGO_DEZ_2026.xlsx
Rodar da raiz:  py update_cards_pipefy/geradores/relatorio_pericias_agendadas.py
"""
import sys, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import _paths
from lib import api
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── configuração ──────────────────────────────────────────────
PIPE_ID   = "305779737"                     # PERÍCIA MÉDICA ADM
ANO       = 2026
MESES     = {8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"}
FASE_TERMINAL = "CONCLUÍDO"                 # já encerrado → fora do worklist

F_VENC = "data_de_vencimento_para_organizar_os_cards"   # DATA E HORA DA PERÍCIA MÉDICA (VENC)
F_NOME = "nome_do_benefici_rio"
F_PROC = "n_mero_do_processo_administrativo"

def nz(v): return (str(v).strip() if v is not None else "")

# ── pull ao vivo ──────────────────────────────────────────────
Q = """query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
  pageInfo{hasNextPage endCursor} edges{node{ id title current_phase{name} fields{ value field{id} } }} } }"""

def pull(pid):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after})
        if "errors" in r: raise RuntimeError(r["errors"])
        d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]
            fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            out.append({
                "phase": nz((n.get("current_phase") or {}).get("name")),
                "nome":  nz(fm.get(F_NOME)) or nz(n.get("title")),
                "proc":  nz(fm.get(F_PROC)),
                "venc":  nz(fm.get(F_VENC)),
            })
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

def parse_dt(v):
    """'DD/MM/YYYY HH:MM' -> (ano, mes, dia, hora, min) ou None."""
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})(?:\s+(\d{2}):(\d{2}))?", v or "")
    if not m: return None
    d, mo, y, hh, mm = m.groups()
    return (int(y), int(mo), int(d), int(hh or 0), int(mm or 0))

print("Puxando cards de PERÍCIA MÉDICA ADM (ao vivo)...")
cards = pull(PIPE_ID)
print(f"  Total de cards no pipe: {len(cards)}")

# ── filtrar Ago–Dez 2026, excluir CONCLUÍDO ───────────────────
por_mes = {m: [] for m in MESES}
for c in cards:
    dt = parse_dt(c["venc"])
    if not dt: continue
    y, mo = dt[0], dt[1]
    if y != ANO or mo not in MESES: continue
    if c["phase"].strip().upper() == FASE_TERMINAL: continue
    por_mes[mo].append({**c, "_dt": dt})

for m in MESES:
    por_mes[m].sort(key=lambda c: c["_dt"])   # ordena por data/hora

total = sum(len(v) for v in por_mes.values())
print(f"\nPerícias agendadas (ativas) Ago–Dez {ANO}: {total}")
for m in MESES:
    print(f"  {MESES[m]:>10}: {len(por_mes[m])}")

# ── estilos (padrão Teles & Costa) ────────────────────────────
NAVY="0D0E2A"; GOLD="F5A623"; WHITE="FFFFFF"; BLUE="1E3A5F"; ALT="F8FAFC"
INK="1E293B"; PROC_BG="FEF9C3"; FASE_FG="1D4ED8"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(h, sz=9, bold=False): return Font(color=h, size=sz, bold=bold, name="Calibri")
def brd():
    s=Side(style='thin', color="CBD5E1"); return Border(left=s,right=s,top=s,bottom=s)
def aln_c(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def aln_l(): return Alignment(horizontal='left',   vertical='center', wrap_text=True)

COLS = [
    ("#",                                 6),
    ("Nome do Beneficiário",             40),
    ("Data da Perícia",                  20),
    ("Nº do Processo Administrativo",    26),
    ("Fase Atual",                       34),
    ("Verificado? (data alterada)",      26),   # coluna de trabalho p/ a equipe
]

def data_fmt(dt):
    y,mo,d,hh,mm = dt
    return f"{d:02d}/{mo:02d}/{y} {hh:02d}:{mm:02d}" if (hh or mm) else f"{d:02d}/{mo:02d}/{y}"

wb = Workbook(); wb.remove(wb.active)

for m in MESES:
    rows = por_mes[m]
    ws = wb.create_sheet(MESES[m])
    ncol = len(COLS); nc = get_column_letter(ncol)

    ws.merge_cells(f"A1:{nc}1")
    ws["A1"] = f"PERÍCIAS MÉDICAS AGENDADAS — {MESES[m].upper()}/{ANO}"
    ws["A1"].font = Font(color=WHITE, size=14, bold=True, name="Calibri")
    ws["A1"].fill = fill(NAVY); ws["A1"].alignment = aln_c()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{nc}2")
    ws["A2"] = (f"{len(rows)} perícia(s) agendada(s)  |  Verificar caso a caso se a data foi "
                f"antecipada pelo INSS  |  Teles & Costa")
    ws["A2"].font = Font(color=GOLD, size=9, bold=True, name="Calibri")
    ws["A2"].fill = fill("161838"); ws["A2"].alignment = aln_c()
    ws.row_dimensions[2].height = 16

    for ci, (hdr, cw) in enumerate(COLS, start=1):
        cell = ws.cell(3, ci, hdr)
        cell.font = fnt(WHITE, 10, True); cell.fill = fill(BLUE)
        cell.alignment = aln_c(); cell.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[3].height = 24

    for idx, r in enumerate(rows):
        row_n = idx + 4
        bg = ALT if idx % 2 == 0 else WHITE
        vals = [idx+1, r["nome"], data_fmt(r["_dt"]), r["proc"], r["phase"], ""]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row_n, ci, val)
            cell.border = brd(); cell.font = fnt(INK, 9)
            cell.alignment = aln_l() if ci in (2, 5) else aln_c()
            cell.fill = fill(bg)
        # destaque nº do processo (amarelo claro) e fase (azul)
        ws.cell(row_n, 4).fill = fill(PROC_BG)
        ws.cell(row_n, 5).font = fnt(FASE_FG, 9, True)
        ws.row_dimensions[row_n].height = 18

    last = 3 + len(rows)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{nc}{last if last>=3 else 3}"

out = _paths.out("pericias", f"PERICIAS_AGENDADAS_AGO_DEZ_{ANO}.xlsx")
wb.save(out)
print(f"\nSalvo: {out}")
