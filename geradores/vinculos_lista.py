# -*- coding: utf-8 -*-
"""Vínculos (terceiro/fundo) APENAS dos casos do arquivo 'Livres - Pipefy.xlsx'.
Casamento por NOME contra o Pipefy ao vivo. Saída: ../Livres_Pipefy_VALIDADO.xlsx"""
import sys, re, os, unicodedata
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
ARQ=_paths.src("Livres - Pipefy.xlsx")
SAIDA=_paths.out("validacoes","Livres_Pipefy_VALIDADO.xlsx")

FID={"ADM":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio_1","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "JUD":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interassado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "FIN":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c
def norm(s):
    s=unicodedata.normalize("NFKD",str(s or "")).encode("ascii","ignore").decode().upper()
    return re.sub(r"\s+"," ",s).strip()

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=FID[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
cards=pull(ADM,"ADM")+pull(JUD,"JUD")+pull(FIN,"FIN")
by_nome=defaultdict(lambda:{"ADM":[],"JUD":[],"FIN":[]})
for c in cards: by_nome[norm(c["nome"])][c["kind"]].append(c)

# nomes do arquivo
wb=openpyxl.load_workbook(ARQ,data_only=True); ws=wb.active
nomes=[r[1] for r in ws.iter_rows(min_row=2,values_only=True) if r[1]]
print(f"Casos no arquivo: {len(nomes)}")

def pick(cards,f):
    for c in cards:
        if c[f]: return c[f]
    return ""

registros=[]
for nome in nomes:
    d=by_nome.get(norm(nome),{"ADM":[],"JUD":[],"FIN":[]})
    todos=d["ADM"]+d["JUD"]+d["FIN"]
    cpf=""
    for c in todos:
        if c["cpf"]: cpf=c["cpf"]; break
    terc_adm=pick(d["ADM"],"terc"); terc_jud=pick(d["JUD"],"terc"); terc_fin=pick(d["FIN"],"terc")
    tvals=[t for t in (terc_adm,terc_jud,terc_fin) if t]
    terc=Counter(tvals).most_common(1)[0][0] if tvals else ""
    registros.append({
        "Nome": nome, "CPF": fcpf(cpf) if cpf else "(não encontrado)",
        "Terceiro Interessado": terc,
        "Terceiro (ADM)": terc_adm, "Terceiro (JUD)": terc_jud, "Terceiro (FIN)": terc_fin,
        "Fundo Investidor": pick(todos,"inv"), "Fundo Especial": pick(todos,"esp"),
        "Fase ADM": pick(d["ADM"],"phase"), "Fase JUD": pick(d["JUD"],"phase"),
        "Pipes": "+".join(k for k in ("ADM","JUD","FIN") if d[k]) or "(nenhum)",
    })

# ── Excel ──
out=openpyxl.Workbook(); wsx=out.active; wsx.title="Vínculos"
headers=["Nome","CPF","Terceiro Interessado","Terceiro (ADM)","Terceiro (JUD)","Terceiro (FIN)",
         "Fundo Investidor","Fundo Especial","Fase ADM","Fase JUD","Pipes"]
wsx.append(headers)
for r in registros: wsx.append([r[h] for h in headers])
font=Font(bold=True,color="FFFFFF"); thin=Side(style="thin",color="BFBFBF")
cf={1:"1F4E78",2:"1F4E78",3:"7C3AED",4:"7C3AED",5:"7C3AED",6:"7C3AED",7:"548235",8:"548235",9:"9E480E",10:"9E480E",11:"404040"}
for i,h in enumerate(headers,1):
    c=wsx.cell(1,i); c.fill=PatternFill("solid",fgColor=cf.get(i,"404040")); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=thin)
wsx.freeze_panes="A2"; wsx.row_dimensions[1].height=30
for i,w in enumerate([32,15,28,22,28,20,20,16,22,26,12],1):
    wsx.column_dimensions[get_column_letter(i)].width=w
out.save(SAIDA)
print("Salvo:",os.path.abspath(SAIDA))
print(f"\n{'NOME':34} {'TERCEIRO':34} FUNDO ESP")
for r in registros:
    print(f"  {r['Nome'][:32]:32} {(r['Terceiro Interessado'] or '-')[:32]:32} {r['Fundo Especial'] or '-'}")
