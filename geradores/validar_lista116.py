# -*- coding: utf-8 -*-
"""Valida a lista TELES-116 contra o Pipefy. Casa por NOME (sem CPF).
Livre = 3 vínculos vazios em ADM+JUD+FIN + não no FIN + em andamento.
Saída: relatorios/validacoes/TELES_116_VALIDADO.xlsx"""
import sys, re, unicodedata
from difflib import SequenceMatcher
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
SAIDA=_paths.out("validacoes","TELES_116_VALIDADO.xlsx")
FID={"ADM":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio_1","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "JUD":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interassado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "FIN":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS","PROCEDENTES"}

NOMES=["ÁGATHA FRANCISCA DE CAMARGO","PEDRO YAN NASCIMENTO SILVA","RALPH LEVY LEMOS TEIXEIRA","ALLAN GABRIEL SILVA SOUZA",
"RUAN GUILHERME LEMOS TEIXEIRA","ANA LAURA DAS GRAÇAS CAMARGO","MIGUEL LUCAS SANTOS MELO","RUBENS MANOEL FERNANDES SEVERINO",
"ANNA KEVELLYN TOMÁS DA SILVA PIRES","ANTONIO HELIO LOPES BORGES","DAVI LUCAS DOS SANTOS GAMA","DHONATA ALEXANDRE DAMASCENO LIMA",
"GABRIEL FERREIRA SANTOS","HARLLEY CALEB PEREIRA VALERIO","MIGUEL LEVY FERNANDES DA SILVA","SARAH BEATRIZ PONCIANO DA SILVA",
"THALYTA VICTORIA SOUZA DA SILVA","ANA CLARA DAS GRAÇAS CAMARGO","ANA KIARA DIAS ALVES","ELOAH VICTORIA MARTINS CAMPOS",
"DAVYLA MELODY QUEIROZ DE MORAIS","FRANCISCO GAEL SILVA SOUSA","FRANCISCO EMANUEL LIMA SOUSA","MARYA KLARA DA FONSECA PEREIRA",
"GABRIEL DE SOUZA CARVALHO","LÍVIA MARIA WEBER MERSONI","ARTHUR GABRIEL BRITO DE ANDRADE","ISAAC GOMES DUTRA",
"LUCAS MIGUEL DE SOUSA QUEIROZ BEZERRA","BRENNO DA SILVA LOPES","DAVID SECUNDINO DE ANDRADE","DAVI LUIZ LOPES BORGES",
"HENRIQUE FERNANDO GOMES MOTA","PAULO MIGUEL RODRIGUES SOUZA","GAEL COSTA MAIA","PEDRO LUCAS SANTOS SILVA","MANOELA DA SILVA",
"NICOLAS MIGUEL DA SILVA FRAZÃO","MIGUEL DA COSTA","ANA BEATRIZ MENESES RODRIGUES","MARIA VICTORIA MENESES RODRIGUES",
"ARTHUR NICOLLAS DIAS GOMES","JHEMILLY VICTORYA ALCANTARA","DANIEL FELIPE DE SOUZA ROSA","HELENA EMANUELLY MEDEIROS DA SILVA",
"LAURA VITORIA DE SOUSA RODRIGUES","HEYTOR AQUINO DE CARVALHO","ANA BEATRIZ VITÓRIA DE JESUS SOUSA","HEITOR COSTA BUENO",
"MIGUEL FERNANDES DA SILVA PEREIRA","EMANUELY STRAPAICCI CASTRO","DANIEL STRAPAICCI FREITAS","BRAYAN SAMUEL DA SILVA COSTA",
"ANTHONY GAEL CORREIA DE ARAÚJO","WENDEL SABINO DA SILVA","KAUANNE VITÓRIA FEIJÓ PEREIRA SILVA","CRISTIANO RODRIGUES DE SOUZA",
"KIMBERLY DA SILVA LARRÉ","AGATHA JAMBEIRO PEREIRA","ADAM LEVY ALVES RODRIGUES","ISAQUE CHRISTOPHER DE SOUSA PEROBA",
"MATIAS NOGUEIRA QUEIROZ","LAURA DE ALMEIDA","ANTHONY MENDONÇA","EMANUELLA VITÓRIA PEREIRA DE BRITO","ARTHUR ALVES DE QUEIROZ",
"LUNA CONCEIÇÃO FRANÇA","FRANCISCO JOSE ALVES ROCHA REGO","ARTHUR MATHEUS SALES HELOTERIO","BRAYAN DA SILVA GONZALEZ",
"ANTONELLA BATISTA DE SÁ","ANTHONNY ISRAEL PEREIRA SILVA","KAIKY RYAN DOS SANTOS CHAGAS","THÉO ALEN DEL MONTE",
"RHAVY LUCCAS DA SILVA DE ARAUJO","MARIA ALICE PEREIRA GUIMARÃES","HENRIQUE LACERDA OLIVEIRA","LUAN GABRIEL OLIVEIRA LOPES",
"ALLANE KAROLINY COSTA FARIA","ANTÔNIO GABRIEL FARIAS MOITA","HILARY ALMEIDA DA SILVA","MARIA FLOR SANTOS DA SILVA",
"KIRIA GABRIELLY SOUZA BISPO CALMON","YASMIM VITÓRIA DE SOUSA LIMA","WERLLY SEVERINO RODRIGUES DE ARRUDA",
"KRISTEN BEATRIZ VIEIRA DE LIMA","MILENE YASMIM BARROSO DE SOUZA","MARIA CLARA ALVES SILVA","KEMILLY LUNA SIQUEIRA DA COSTA",
"GAEL CRISTHIAM AKERLEY","JULIO CESAR RODRIGUES SANTANA","VITÓRIA EMANUELLY GONÇALVES DOS SANTOS",
"RAIMUNDA STHERFANY SOUSA DO NASCIMENTO","NICOLAS JHOSEF PEREIRA DE JESUS","NOAH DE OLIVEIRA MARTINS","URIEL ALVES PINTO",
"ANA LIZ LIMA ALVES","MARIA VITÓRIA SILVA FERREIRA","SARAH INGRID NASCIMENTO OLIVEIRA","ISABELA VITORIA SANTOS MENDES",
"AGHATA DO NASCIMENTO PEREIRA","THEO BERNARDO DA SILVA","LORENA EMANUELLE ALVES BRANDÃO","QUEYLLA ISABELI DA SILVA DE SOUZA",
"MIKAELLY RODRIGUES DA SILVA","TAYLLA VICTORIA CABRAL DA SILVA","WILLIAM GABRIEL VILELA CAETANO","ANA JULIA SOARES DA COSTA",
"MARJORIE ALLYSSA SILVA DE LIMA","KAYNAN DOS SANTOS CELLINE","MARIA LUIZA SANTOS DE BRITO","ENZO GABRIEL SANTOS DE BRITO",
"ENZO GABRIEL ARAUJO DA SILVA","ADRYAN HENRYCK OLIVEIRA DO COUTO CORREA","CALLEB ELIAS GONÇALVES CARDOSO",
"ISABELA MARIA DE CASTRO RIBEIRO"]

def norm(s):
    s=unicodedata.normalize("NFKD",str(s or "")).encode("ascii","ignore").decode().upper()
    return re.sub(r'\s+',' ',s).strip()
def nz(v): return (str(v).strip() if v is not None else "")
def dig(v): return re.sub(r"\D","",str(v or ""))
def c11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=FID[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":c11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
cards=pull(ADM,"ADM")+pull(JUD,"JUD")+pull(FIN,"FIN")
by_cpf=defaultdict(lambda:{"ADM":[],"JUD":[],"FIN":[],"nome":""})
name2cpf=defaultdict(set); all_names=set()
for c in cards:
    if c["cpf"]:
        by_cpf[c["cpf"]][c["kind"]].append(c); by_cpf[c["cpf"]]["nome"]=c["nome"] or by_cpf[c["cpf"]]["nome"]
        name2cpf[norm(c["nome"])].add(c["cpf"])
    all_names.add(norm(c["nome"]))

def active(c):
    term=ADM_TERM if c["kind"]=="ADM" else (JUD_TERM if c["kind"]=="JUD" else set())
    return c["phase"].upper() not in term
def avalia(cpf):
    d=by_cpf[cpf]; todos=d["ADM"]+d["JUD"]+d["FIN"]; vinc=[]
    for c in todos:
        v=c["terc"] or c["inv"] or c["esp"]
        if v: vinc.append(f'{c["kind"]}={v}')
    in_fin=bool(d["FIN"]); ativo=any(active(c) for c in d["ADM"]+d["JUD"]); mot=[]
    if vinc: mot.append("VÍNCULO: "+"; ".join(dict.fromkeys(vinc)))
    if in_fin: mot.append("NO FINANCEIRO")
    if not ativo: mot.append("sem card ativo")
    fase=next((c["phase"] for c in d["ADM"]),"") or next((c["phase"] for c in d["JUD"]),"")
    return ("LIVRE" if not mot else "NÃO LIVRE"), " | ".join(mot), fase

regs=[]
for nome in NOMES:
    nn=norm(nome); cpfs=name2cpf.get(nn)
    if cpfs:
        for cpf in cpfs:
            ver,mot,fase=avalia(cpf); regs.append([nome,fcpf(cpf),"EXATO",ver,mot,fase,by_cpf[cpf]["nome"]])
    else:
        best=max(all_names,key=lambda x:SequenceMatcher(None,nn,x).ratio()) if all_names else ""
        rt=SequenceMatcher(None,nn,best).ratio() if best else 0
        if rt>=0.88:
            for cpf in name2cpf.get(best,[]):
                ver,mot,fase=avalia(cpf); regs.append([nome,fcpf(cpf),f"APROX {rt:.0%}",ver,mot,fase,by_cpf[cpf]["nome"]])
        else:
            regs.append([nome,"","NÃO ENCONTRADO","—","sem card no Pipefy (nome não localizado)","",
                         f"+próximo: {best.title()} ({rt:.0%})" if best else ""])

out=openpyxl.Workbook(); ws=out.active; ws.title="Validação 116"
headers=["Nome (lista)","CPF","Match","Veredito","Motivo","Fase (Pipefy)","Nome no Pipefy"]
ws.append(headers); font=Font(bold=True,color="FFFFFF")
for i,h in enumerate(headers,1):
    c=ws.cell(1,i); c.fill=PatternFill("solid",fgColor="1F4E78"); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=Side(style="thin",color="BFBFBF"))
BG={"LIVRE":"DCFCE7","NÃO LIVRE":"FEE2E2"}
for r in regs:
    ws.append(r); rr=ws.max_row
    ws.cell(rr,4).fill=PatternFill("solid",fgColor=BG.get(r[3],"F1F5F9"))
    if r[2].startswith("APROX"): ws.cell(rr,3).font=Font(bold=True,color="9A3412")
ws.freeze_panes="A2"
for i,w in enumerate([34,15,14,14,46,24,30],1): ws.column_dimensions[get_column_letter(i)].width=w
try: out.save(SAIDA); print("Salvo:",SAIDA)
except PermissionError:
    alt=SAIDA.replace(".xlsx","_v2.xlsx"); out.save(alt); print("[!] travado — salvo:",alt)

livre=sum(1 for r in regs if r[3]=="LIVRE"); nl=sum(1 for r in regs if r[3]=="NÃO LIVRE")
ne=sum(1 for r in regs if r[2]=="NÃO ENCONTRADO")
print(f"\nEntradas: {len(regs)} (nomes: {len(NOMES)}) | LIVRES: {livre} | NÃO LIVRES: {nl} | NÃO ENCONTRADOS: {ne}")
print("\n== NÃO ENCONTRADOS ==")
for r in regs:
    if r[2]=="NÃO ENCONTRADO": print("  -",r[0],"  ",r[6])
print("\n== LIVRES ==")
for r in regs:
    if r[3]=="LIVRE": print("  -",r[0],r[1],"|",r[5])
