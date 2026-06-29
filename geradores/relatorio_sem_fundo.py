# -*- coding: utf-8 -*-
"""Processos SEM FUNDO disponíveis no Pipefy:
 - 3 vínculos vazios (Terceiro Interessado + Fundo Investidor + Fundo Especial) em ADM+JUD+FIN
 - NÃO está no financeiro (FIN)
 - SEM decisão pela etapa atual: em JUD = sentença OU recurso inominado; senão = resultado ADM
1 linha por CPF, Situação como 1ª coluna. Saída: relatorios/processos/PROCESSOS_SEM_FUNDO.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
SAIDA=_paths.out("processos","PROCESSOS_SEM_FUNDO.xlsx")

F={"ADM":{"cpf":"cpf_do_benefici_rio_1","nome":"nome_do_benefici_rio","terc":"terceiro_interessado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor",
          "prot":"n_mero_do_processo_administrativo","data":"data_do_protocolo","res":"resultado_do_pedido"},
   "JUD":{"cpf":"cpf_do_benefici_rio","nome":"nome_do_benefici_rio","terc":"terceiro_interassado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor",
          "prot":"n_mero_do_processo","data":"data_do_protocolo","res":"resultado_do_processo","rec":"resultado_do_recurso"},
   "FIN":{"cpf":"cpf_do_benefici_rio","nome":"nome_do_benefici_rio","terc":"terceiro_interessado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS","PROCEDENTES"}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c
def fdata(v):
    m=re.match(r"(\d{4})-(\d{2})-(\d{2})",nz(v)); return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else nz(v)

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=F[kind]; rec=fm.get(fi["rec"]) if fi.get("rec") else None
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"])),
                "prot":nz(fm.get(fi["prot"])) if fi.get("prot") else "",
                "data":nz(fm.get(fi["data"])) if fi.get("data") else "",
                "res":nz(fm.get(fi["res"])) if fi.get("res") else "","rec":nz(rec)})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
adm_by=defaultdict(list); jud_by=defaultdict(list); fin_by=defaultdict(list)
for c in pull(ADM,"ADM"):
    if c["cpf"]: adm_by[c["cpf"]].append(c)
for c in pull(JUD,"JUD"):
    if c["cpf"]: jud_by[c["cpf"]].append(c)
for c in pull(FIN,"FIN"):
    if c["cpf"]: fin_by[c["cpf"]].append(c)
print(f"  ADM {len(adm_by)} | JUD {len(jud_by)} | FIN {len(fin_by)} (CPFs distintos)")

def is_active(c):
    term=ADM_TERM if c["kind"]=="ADM" else (JUD_TERM if c["kind"]=="JUD" else set())
    return c["phase"].upper() not in term
def pick(cards):
    if not cards: return None
    ativos=[c for c in cards if is_active(c)]
    pool=ativos if ativos else cards
    return max(pool,key=lambda c:(c.get("data",""),c.get("prot","")))

def tem_vinculo(cpf):
    for c in adm_by.get(cpf,[])+jud_by.get(cpf,[])+fin_by.get(cpf,[]):
        if c["terc"] or c["inv"] or c["esp"]: return True
    return False

registros=[]
cpfs=set(adm_by)|set(jud_by)|set(fin_by)
for cpf in cpfs:
    if tem_vinculo(cpf): continue   # tem fundo/terceiro -> fora
    if cpf in fin_by:    continue   # remove casos no financeiro
    jud_cards=jud_by.get(cpf,[]); adm_cards=adm_by.get(cpf,[])
    # decisão pela etapa atual: em JUD = sentença OU recurso; senão = resultado ADM
    if jud_cards:
        decidido = any((c["res"] or c["rec"]) for c in jud_cards)
    else:
        decidido = any(c["res"] for c in adm_cards)
    if decidido: continue           # remove casos com decisão
    a=pick(adm_cards); j=pick(jud_cards)
    in_fin=False
    nome=(a or j or {}).get("nome") or ""
    jph=j["phase"] if j else ""
    # situação
    if j and is_active(j): situ="JUD"; fase_atual=j["phase"]
    elif a and is_active(a): situ="ADM"; fase_atual=a["phase"]
    else: situ="ENCERRADO"; fase_atual=(j["phase"] if j else (a["phase"] if a else ""))
    if fase_atual.upper()=="DESCARTADOS": continue   # regra: não trazer cards descartados
    registros.append({
        "Situação":situ,"Fase Atual":fase_atual,
        "Nome":nome,"CPF":fcpf(cpf),
        "Terceiro Interessado":"","Fundo Investidor":"","Fundo Especial":"",
        "Nº Protocolo (ADM)":a["prot"] if a else "",
        "Data Protocolo (ADM)":fdata(a["data"]) if a else "",
        "Resultado Administrativo":a["res"] if a else "",
        "Nº Protocolo (JUD)":j["prot"] if j else "",
        "Data Protocolo (JUD)":fdata(j["data"]) if j else "",
        "Fase 8 - Aguardando Julgamento (1ª inst.)":"Sim" if jph.upper().startswith("08") else "",
        "Fase 9 - Em Grau Recursal (2ª inst.)":"Sim" if jph.upper().startswith("09") else "",
        "Resultado Recurso Inominado":j["rec"] if j else "",
    })
registros.sort(key=lambda x:(x["Situação"],x["Fase Atual"],str(x["Nome"])))
print(f"SEM fundo, SEM financeiro e SEM decisão (por etapa): {len(registros)}")
print("Por situação:", dict(Counter(r["Situação"] for r in registros)))

# ── Validação: confirmar 3 campos de fundo vazios em ADM e JUD ──
def _vinc_em(cpf):
    bad=[]
    for kind,store in (("ADM",adm_by),("JUD",jud_by)):
        for c in store.get(cpf,[]):
            v=c["terc"] or c["inv"] or c["esp"]
            if v: bad.append(f"{kind}={v}")
    return bad
viol=[(r["Nome"],r["CPF"],_vinc_em(re.sub(r"\D","",r["CPF"]))) for r in registros]
viol=[x for x in viol if x[2]]
print(f"VALIDAÇÃO (3 campos vazios em ADM+JUD): {len(registros)-len(viol)}/{len(registros)} OK | violações: {len(viol)}")
for n,c,b in viol[:30]: print("  [X]",n,c,b)

# ── Excel ── (Situação + Fase Atual como 1ªs colunas)
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Sem Fundo"
headers=["Situação","Fase Atual","Nome","CPF","Terceiro Interessado","Fundo Investidor","Fundo Especial",
         "Nº Protocolo (ADM)","Data Protocolo (ADM)","Resultado Administrativo",
         "Nº Protocolo (JUD)","Data Protocolo (JUD)",
         "Fase 8 - Aguardando Julgamento (1ª inst.)","Fase 9 - Em Grau Recursal (2ª inst.)","Resultado Recurso Inominado"]
ws.append(headers)
for r in registros: ws.append([r.get(h) for h in headers])
font=Font(bold=True,color="FFFFFF"); thin=Side(style="thin",color="BFBFBF")
cf={1:"6B21A8",2:"6B21A8"}            # Situação + Fase Atual = roxo
for i in range(3,8):  cf[i]="1F4E78"  # identificação/vínculos = azul
for i in range(8,11): cf[i]="548235"  # ADM = verde
for i in range(11,16):cf[i]="9E480E"  # JUD = laranja
for i,h in enumerate(headers,1):
    c=ws.cell(1,i); c.fill=PatternFill("solid",fgColor=cf.get(i,"404040")); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=thin)
ws.freeze_panes="C2"; ws.row_dimensions[1].height=42; ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{len(registros)+1}"
for i,w in enumerate([13,28,34,15,18,16,15,20,16,24,20,16,18,18,22],1):
    ws.column_dimensions[get_column_letter(i)].width=w

ws2=wb.create_sheet("Resumo")
ws2.append(["Métrica","Valor"])
for c in ws2[1]: c.font=font; c.fill=PatternFill("solid",fgColor="1F4E78")
ws2.append(["Total sem fundo (3 campos vazios)",len(registros)])
for k,v in Counter(r["Situação"] for r in registros).most_common(): ws2.append([f"  Situação: {k}",v])
ws2.append(["Com protocolo ADM",sum(1 for r in registros if r['Nº Protocolo (ADM)'])])
ws2.append(["Com protocolo JUD",sum(1 for r in registros if r['Nº Protocolo (JUD)'])])
ws2.append([])
ws2.append(["Por Fase Atual",""])
for k,v in Counter(r["Fase Atual"] or "(sem fase)" for r in registros).most_common(): ws2.append([k,v])
ws2.column_dimensions["A"].width=40; ws2.column_dimensions["B"].width=12

try:
    wb.save(SAIDA); print("Salvo:",os.path.abspath(SAIDA))
except PermissionError:
    alt=SAIDA.replace(".xlsx","_v2.xlsx"); wb.save(alt); print("[!] travado — salvo como:",os.path.abspath(alt))
