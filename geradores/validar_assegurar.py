# -*- coding: utf-8 -*-
"""Valida 'Protocolos - Assegurar (2).xlsx' (aba Processos) — estão livres?
Livre = 3 vínculos vazios (Terceiro/Fundo Inv/Fundo Esp) em ADM+JUD+FIN + NÃO no financeiro + em andamento.
Casamento por CPF. Saída: ../Assegurar_VALIDADO.xlsx"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
ARQ=_paths.ucp("Protocolos - Assegurar (2).xlsx")
SAIDA=_paths.out("validacoes","Assegurar_VALIDADO.xlsx")

FID={"ADM":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio_1","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "JUD":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interassado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"},
     "FIN":{"nome":"nome_do_benefici_rio","cpf":"cpf_do_benefici_rio","terc":"terceiro_interessado","inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS","PROCEDENTES"}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=FID[kind]
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"]))})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
by=defaultdict(lambda:{"ADM":[],"JUD":[],"FIN":[]})
for c in pull(ADM,"ADM")+pull(JUD,"JUD")+pull(FIN,"FIN"):
    if c["cpf"]: by[c["cpf"]][c["kind"]].append(c)

wb=openpyxl.load_workbook(ARQ,data_only=True); ws=wb["Processos"]
casos=[(r[0],cpf11(r[1])) for r in ws.iter_rows(min_row=2,values_only=True) if r[1]]
print(f"Casos no arquivo: {len(casos)}\n")

def active(c):
    term=ADM_TERM if c["kind"]=="ADM" else (JUD_TERM if c["kind"]=="JUD" else set())
    return c["phase"].upper() not in term

registros=[]; livres=0
for nome,cpf in casos:
    d=by.get(cpf);
    if not d:
        registros.append({"Nome":nome,"CPF":fcpf(cpf),"Veredito":"NÃO ENCONTRADO","Motivo":"CPF sem card no Pipefy",
            "Terceiro (ADM)":"","Terceiro (JUD)":"","Fundo Inv":"","Fundo Esp":"","No Financeiro?":"","Fase ADM":"","Fase JUD":""}); continue
    todos=d["ADM"]+d["JUD"]+d["FIN"]
    terc_adm=next((c["terc"] for c in d["ADM"] if c["terc"]),"")
    terc_jud=next((c["terc"] for c in d["JUD"] if c["terc"]),"")
    terc_fin=next((c["terc"] for c in d["FIN"] if c["terc"]),"")
    inv=next((c["inv"] for c in todos if c["inv"]),"")
    esp=next((c["esp"] for c in todos if c["esp"]),"")
    in_fin=bool(d["FIN"])
    tem_ativo=any(active(c) for c in d["ADM"]+d["JUD"])
    motivos=[]
    vincs=[]
    if terc_adm: vincs.append(f"Terc.ADM={terc_adm}")
    if terc_jud: vincs.append(f"Terc.JUD={terc_jud}")
    if terc_fin: vincs.append(f"Terc.FIN={terc_fin}")
    if inv: vincs.append(f"FundoInv={inv}")
    if esp: vincs.append(f"FundoEsp={esp}")
    if vincs: motivos.append("VÍNCULO: "+"; ".join(vincs))
    if in_fin: motivos.append("ESTÁ NO FINANCEIRO")
    if not tem_ativo: motivos.append("sem card ativo (não em andamento)")
    veredito="LIVRE" if not motivos else "NÃO LIVRE"
    if not motivos: livres+=1
    registros.append({"Nome":nome,"CPF":fcpf(cpf),"Veredito":veredito,"Motivo":" | ".join(motivos),
        "Terceiro (ADM)":terc_adm,"Terceiro (JUD)":terc_jud,"Fundo Inv":inv,"Fundo Esp":esp,
        "No Financeiro?":"Sim" if in_fin else "",
        "Fase ADM":next((c["phase"] for c in d["ADM"]),""),"Fase JUD":next((c["phase"] for c in d["JUD"]),"")})

# Excel
out=openpyxl.Workbook(); wsx=out.active; wsx.title="Validação"
headers=["Nome","CPF","Veredito","Motivo","Terceiro (ADM)","Terceiro (JUD)","Fundo Inv","Fundo Esp","No Financeiro?","Fase ADM","Fase JUD"]
wsx.append(headers)
GREEN="DCFCE7"; RED="FEE2E2"; GRAY="F1F5F9"
for r in registros:
    wsx.append([r[h] for h in headers])
    bg = GREEN if r["Veredito"]=="LIVRE" else (GRAY if r["Veredito"]=="NÃO ENCONTRADO" else RED)
    wsx.cell(wsx.max_row,3).fill=PatternFill("solid",fgColor=bg)
font=Font(bold=True,color="FFFFFF")
for i,h in enumerate(headers,1):
    c=wsx.cell(1,i); c.fill=PatternFill("solid",fgColor="1F4E78"); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=Side(style="thin",color="BFBFBF"))
wsx.freeze_panes="A2"; wsx.row_dimensions[1].height=28
for i,w in enumerate([32,15,14,46,22,26,18,14,13,24,26],1):
    wsx.column_dimensions[get_column_letter(i)].width=w
out.save(SAIDA)

nl=sum(1 for r in registros if r["Veredito"]=="NÃO LIVRE")
ne=sum(1 for r in registros if r["Veredito"]=="NÃO ENCONTRADO")
print(f"LIVRES: {livres} | NÃO LIVRES: {nl} | NÃO ENCONTRADOS: {ne} | total {len(registros)}")
print("Salvo:",os.path.abspath(SAIDA))
print("\nNÃO LIVRES (amostra):")
for r in registros:
    if r["Veredito"]=="NÃO LIVRE":
        print(f"  {r['Nome'][:30]:30} -> {r['Motivo'][:80]}")
