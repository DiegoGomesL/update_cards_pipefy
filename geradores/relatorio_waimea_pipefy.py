# -*- coding: utf-8 -*-
"""Todos os processos do FUNDO WAIMEA que estão no Pipefy, no modelo PROCESSOS_LIVRES.
Waimea = CPFs da carteira (WAIMEA_900_Processos.xlsx) que possuem card no Pipefy (ADM/JUD/FIN).
1 linha por CPF, consolidando ADM+JUD. Saída: ../WAIMEA_PIPEFY.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import _paths
from lib import api

ADM=api.PIPES["ADM"]["id"]; JUD=api.PIPES["JUD"]["id"]; FIN=api.PIPES["FIN"]["id"]
CARTEIRA=_paths.src("WAIMEA_900_Processos.xlsx")
SAIDA=_paths.out("waimea","WAIMEA_PIPEFY.xlsx")

F={"ADM":{"cpf":"cpf_do_benefici_rio_1","nome":"nome_do_benefici_rio","terc":"terceiro_interessado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor",
          "prot":"n_mero_do_processo_administrativo","data":"data_do_protocolo","res":"resultado_do_pedido"},
   "JUD":{"cpf":"cpf_do_benefici_rio","nome":"nome_do_benefici_rio","terc":"terceiro_interassado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor",
          "prot":"n_mero_do_processo","data":"data_do_protocolo","res":"resultado_do_processo","rec":"resultado_do_recurso"},
   "FIN":{"cpf":"cpf_do_benefici_rio","nome":"nome_do_benefici_rio","terc":"terceiro_interessado",
          "inv":"fundo_investidor","esp":"copy_of_fundo_investidor"}}
ADM_TERM={"ENCERRADOS","DESCARTADOS"}; JUD_TERM={"ENCERRADOS","PROCEDENTES"}

def dig(v): return re.sub(r"\D","",str(v or ""))
def cpf11(v):
    d=dig(v); return d.zfill(11) if d and len(d)<=11 else d
def nz(v): return (str(v).strip() if v is not None else "")
def fcpf(c): return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" if len(c)==11 else c
def fdata(v):
    m=re.match(r"(\d{4})-(\d{2})-(\d{2})",nz(v)); return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else nz(v)

# carteira Waimea
wb=openpyxl.load_workbook(CARTEIRA,read_only=True,data_only=True); ws=wb["Processos"]
rows=list(ws.iter_rows(values_only=True)); wb.close()
waimea_nome={}
for r in rows[1:]:
    c=cpf11(r[9])
    if c and len(c)==11: waimea_nome[c]=r[7]  # nome da carteira
print(f"CPFs Waimea na carteira: {len(waimea_nome)}")

Q="""query($pipeId:ID!,$after:String){ allCards(pipeId:$pipeId,first:50,after:$after){
 pageInfo{hasNextPage endCursor} edges{node{ id current_phase{name} fields{ value field{id} } }}}}"""
def pull(pid,kind):
    out=[]; after=None
    while True:
        r=api.execute(Q,{"pipeId":pid,"after":after}); d=r["data"]["allCards"]
        for e in d["edges"]:
            n=e["node"]; fm={(f.get("field") or {}).get("id"):f.get("value") for f in (n.get("fields") or [])}
            fi=F[kind]; rec=fm.get(fi["rec"]) if fi.get("rec") else None
            out.append({"kind":kind,"phase":nz((n.get("current_phase") or {}).get("name")),
                "cpf":cpf11(fm.get(fi["cpf"])),"nome":nz(fm.get(fi["nome"])),
                "terc":nz(fm.get(fi["terc"])),"inv":nz(fm.get(fi["inv"])),"esp":nz(fm.get(fi["esp"])),
                "prot":nz(fm.get(fi["prot"])) if fi.get("prot") else "",
                "data":nz(fm.get(fi["data"])) if fi.get("data") else "",
                "res":nz(fm.get(fi["res"])) if fi.get("res") else "","rec":nz(rec)})
        if not d["pageInfo"]["hasNextPage"]: break
        after=d["pageInfo"]["endCursor"]
    return out

print("Buscando Pipefy...")
adm_by=defaultdict(list); jud_by=defaultdict(list); fin_by=defaultdict(list)
for c in pull(ADM,"ADM"):
    if c["cpf"] in waimea_nome: adm_by[c["cpf"]].append(c)
for c in pull(JUD,"JUD"):
    if c["cpf"] in waimea_nome: jud_by[c["cpf"]].append(c)
for c in pull(FIN,"FIN"):
    if c["cpf"] in waimea_nome: fin_by[c["cpf"]].append(c)

def is_active(c):
    term=ADM_TERM if c["kind"]=="ADM" else (JUD_TERM if c["kind"]=="JUD" else set())
    return c["phase"].upper() not in term
def pick(cards):
    if not cards: return None
    ativos=[c for c in cards if is_active(c)]
    pool=ativos if ativos else cards
    return max(pool,key=lambda c:(c.get("data",""),c.get("prot","")))
def pickv(cpf,field):
    for src in (adm_by,jud_by,fin_by):
        for c in src.get(cpf,[]):
            if c.get(field): return c[field]
    return ""

cpfs=set(adm_by)|set(jud_by)|set(fin_by)
print(f"CPFs Waimea encontrados no Pipefy: {len(cpfs)} (de {len(waimea_nome)})")

registros=[]
for cpf in cpfs:
    a=pick(adm_by.get(cpf,[])); j=pick(jud_by.get(cpf,[]))
    nome=(a or j or pick(fin_by.get(cpf,[])) or {}).get("nome") or waimea_nome.get(cpf,"")
    jph=j["phase"] if j else ""
    in_fin = cpf in fin_by
    registros.append({
        "Nome":nome,"CPF":fcpf(cpf),
        "Terceiro Interessado":pickv(cpf,"terc"),
        "Fundo Investidor":pickv(cpf,"inv"),
        "Fundo Especial":pickv(cpf,"esp"),
        "Fase do card (ADM)":a["phase"] if a else "",
        "Nº Protocolo (ADM)":a["prot"] if a else "",
        "Data Protocolo (ADM)":fdata(a["data"]) if a else "",
        "Resultado Administrativo":a["res"] if a else "",
        "Fase do card (JUD)":jph,
        "Nº Protocolo (JUD)":j["prot"] if j else "",
        "Data Protocolo (JUD)":fdata(j["data"]) if j else "",
        "Fase 8 - Aguardando Julgamento (1ª inst.)":"Sim" if jph.upper().startswith("08") else "",
        "Fase 9 - Em Grau Recursal (2ª inst.)":"Sim" if jph.upper().startswith("09") else "",
        "Resultado Recurso Inominado":j["rec"] if j else "",
        "No Financeiro?":"Sim" if in_fin else "",
    })
registros.sort(key=lambda x:(x["Fase do card (JUD)"],x["Fase do card (ADM)"],str(x["Nome"])))

# ── Excel ──
wbx=openpyxl.Workbook(); wsx=wbx.active; wsx.title="Processos Waimea"
headers=list(registros[0].keys()) if registros else ["Nome","CPF"]
wsx.append(headers)
for r in registros: wsx.append([r.get(h) for h in headers])
font=Font(bold=True,color="FFFFFF"); thin=Side(style="thin",color="BFBFBF")
cf={}
for i in range(1,6):  cf[i]="1F4E78"
for i in range(6,10): cf[i]="548235"
for i in range(10,16):cf[i]="9E480E"
cf[16]="404040"
for i,h in enumerate(headers,1):
    c=wsx.cell(1,i); c.fill=PatternFill("solid",fgColor=cf.get(i,"404040")); c.font=font
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=Border(bottom=thin)
wsx.freeze_panes="A2"; wsx.row_dimensions[1].height=42
for i,w in enumerate([34,15,24,16,15,26,22,16,24,26,22,16,18,18,24,14],1):
    wsx.column_dimensions[get_column_letter(i)].width=w

# Resumo
ws2=wbx.create_sheet("Resumo")
ws2.append(["Métrica","Valor"])
for c in ws2[1]: c.font=font; c.fill=PatternFill("solid",fgColor="1F4E78")
ws2.append(["CPFs Waimea na carteira", len(waimea_nome)])
ws2.append(["Encontrados no Pipefy", len(registros)])
ws2.append(["No Financeiro (FIN)", sum(1 for r in registros if r['No Financeiro?']=='Sim')])
ws2.append(["Com card JUD", sum(1 for r in registros if r['Fase do card (JUD)'])])
ws2.append(["Com protocolo ADM", sum(1 for r in registros if r['Nº Protocolo (ADM)'])])
ws2.append(["Com protocolo JUD", sum(1 for r in registros if r['Nº Protocolo (JUD)'])])
ws2.append([])
ws2.append(["Por Fase ADM",""])
for k,v in Counter(r["Fase do card (ADM)"] or "(sem ADM)" for r in registros).most_common(): ws2.append([k,v])
ws2.append([])
ws2.append(["Por Fundo Especial",""])
for k,v in Counter(r["Fundo Especial"] or "(vazio)" for r in registros).most_common(): ws2.append([k,v])
ws2.column_dimensions["A"].width=40; ws2.column_dimensions["B"].width=12

wbx.save(SAIDA)
print("Salvo:",os.path.abspath(SAIDA))
print(f"\nTotal: {len(registros)} | No FIN: {sum(1 for r in registros if r['No Financeiro?']=='Sim')} | Com JUD: {sum(1 for r in registros if r['Fase do card (JUD)'])}")
